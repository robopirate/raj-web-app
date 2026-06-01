# --- Calendar & Drive Integration (v4.0) ---
CALENDAR_AVAILABLE = False
DRIVE_AVAILABLE = False
try:
    from calendar_integration import CalendarManager
    CALENDAR_AVAILABLE = True
except ImportError:
    pass
try:
    from drive_integration import DriveManager
    DRIVE_AVAILABLE = True
except ImportError:
    pass
# --- End Integration ---

"""
engine.py -- RoboPirate Campaign Engine v3.1
SCHOOL + CSR sequences | Raj as manager | Auto-send | Draft-only replies
"""

import re
import json
import time
import threading
from datetime import datetime, timedelta
from typing import List, Dict, Tuple, Optional
from dataclasses import dataclass
from collections import deque

from db import Database
from gmail import GmailClient

try:
    from smart_importer import SmartImporter
    SMART_IMPORT_AVAILABLE = True
except ImportError:
    SMART_IMPORT_AVAILABLE = False

# Sequences: SCHOOL (private schools) and CSR (corporates)
# Raj is the manager/agent persona, not a sequence
SEQUENCES = {
    "school": {
        "days": [1, 3, 5, 7, 10],
        "template_prefix": "SCHOOL EMAIL ",
        "audience": "private_school",
        "persona": "school",
        "assets": {
            1: {
                "brochure": "https://drive.google.com/file/d/1vRMeFM22aajc5zfiYhqaev34UVQ87zyU/view",
                "video_wsl": "https://drive.google.com/file/d/1KPrC2IpdooxazGJiyVe79JgyWlJbOxzu/view",
                "video_abp": "https://youtu.be/FJ2_W53WjmA",
                "video_ig": "https://www.instagram.com/reel/DMe2HzqofAk/"
            },
            3: {
                "video_abp": "https://youtu.be/FJ2_W53WjmA",
                "video_ig": "https://www.instagram.com/reel/DMe2HzqofAk/"
            },
            5: {
                "report_vbv": "https://drive.google.com/file/d/1d7EEtC8YitbSj7U6ivHf_6WtUGuylT-B/view",
                "video_star": "https://youtube.com/watch?v=iziKPBSfGKU",
                "folder_vbv": "https://drive.google.com/drive/folders/1tWu3zrH0zIjJbkfS3hX0tnKXQY-9HTgN",
                "video_ig": "https://www.instagram.com/reel/DMe2HzqofAk/"
            },
            7: {
                "profile": "https://drive.google.com/file/d/1g9JJ4_VO_28QKYD7iVVDJZcv9l4uRbZu/view",
                "video_abp": "https://youtu.be/FJ2_W53WjmA",
                "video_star": "https://youtube.com/watch?v=iziKPBSfGKU",
                "video_ig": "https://www.instagram.com/reel/DMe2HzqofAk/"
            },
            10: {
                "plans": "https://drive.google.com/file/d/1vRMeFM22aajc5zfiYhqaev34UVQ87zyU/view",
                "video_ig": "https://www.instagram.com/reel/DMe2HzqofAk/"
            }
        }
    },
    "csr": {
        "days": [1, 3, 5, 7, 10],
        "template_prefix": "CSR EMAIL ",
        "audience": "csr",
        "persona": "csr",
        "assets": {
            1: {
                "report_sangli1": "https://drive.google.com/file/d/1HpNdnamA2k3H0xkKr58STEKMNu5RgHPx/view",
                "video_abp": "https://youtu.be/FJ2_W53WjmA",
                "video_sangli": "https://drive.google.com/file/d/1MUlsC87vRbhFaoW0XcX146WBLKYBk448/view",
                "video_ig": "https://www.instagram.com/reel/DMe2HzqofAk/"
            },
            3: {
                "report_sangli1": "https://drive.google.com/file/d/1HpNdnamA2k3H0xkKr58STEKMNu5RgHPx/view",
                "brochure": "https://drive.google.com/file/d/1vRMeFM22aajc5zfiYhqaev34UVQ87zyU/view",
                "video_ig": "https://www.instagram.com/reel/DMe2HzqofAk/"
            },
            5: {
                "report_sangli2": "https://drive.google.com/file/d/1pKSm1WPlPk-we4aC-uhqxEy8w-BYygSN/view",
                "report_vbv": "https://drive.google.com/file/d/1d7EEtC8YitbSj7U6ivHf_6WtUGuylT-B/view",
                "video_star": "https://youtube.com/watch?v=iziKPBSfGKU",
                "folder_sangli": "https://drive.google.com/drive/folders/15sc5iOIKTBZyenb2rCpGVAK1lExcG5BC",
                "video_ig": "https://www.instagram.com/reel/DMe2HzqofAk/"
            },
            7: {
                "plans": "https://drive.google.com/file/d/1vRMeFM22aajc5zfiYhqaev34UVQ87zyU/view",
                "video_wsl": "https://drive.google.com/file/d/1KPrC2IpdooxazGJiyVe79JgyWlJbOxzu/view",
                "video_abp": "https://youtu.be/FJ2_W53WjmA",
                "video_sangli": "https://drive.google.com/file/d/1MUlsC87vRbhFaoW0XcX146WBLKYBk448/view",
                "video_ig": "https://www.instagram.com/reel/DMe2HzqofAk/"
            },
            10: {
                "profile": "https://drive.google.com/file/d/1g9JJ4_VO_28QKYD7iVVDJZcv9l4uRbZu/view",
                "kits": "https://drive.google.com/file/d/1cvi4p8IHgx1MekanVRHN3Fo4Lk9vbubX/view",
                "video_ig": "https://www.instagram.com/reel/DMe2HzqofAk/"
            }
        }
    }
}

EMAIL_NUM_TO_DAY = {1: 1, 2: 3, 3: 5, 4: 7, 5: 10}
DAY_TO_EMAIL_NUM = {1: 1, 3: 2, 5: 3, 7: 4, 10: 5}

SEND_DELAY = 45
BOUNCE_INTERVAL = 6
REPLY_INTERVAL = 60
EMERGENCY_INTERVAL = 15
EOD_HOUR = 19
MORNING_HOUR = 8

HTML_TEMPLATE = """<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0"></head>
<body style="margin:0;padding:0;font-family:'Segoe UI',Arial,sans-serif;background:#f5f5f5;">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f5f5f5;">
<tr><td align="center" style="padding:20px 0;">
<table width="600" cellpadding="0" cellspacing="0" style="background:#ffffff;border-radius:8px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,0.1);">
<tr><td style="background:linear-gradient(135deg,#0D9B8A,#0BC5B0);padding:20px;text-align:center;">
<img src="https://robopirate.in/assets/logo-CCM9tiYQ.png" alt="RoboPirate" style="height:50px;margin-bottom:10px;">
<h1 style="color:#ffffff;margin:0;font-size:22px;">{title}</h1>
</td></tr>
<tr><td style="padding:30px;color:#333;line-height:1.6;">
{content}
</td></tr>
<tr><td style="background:#1A6B6B;padding:20px;text-align:center;">
<p style="color:#ffffff;margin:0 0 10px;font-size:14px;">
<strong style="color:#F5A623;">85+</strong> Labs | 
<strong style="color:#F5A623;">65K+</strong> Students | 
<strong style="color:#F5A623;">6</strong> States
</p>
<a href="https://wa.me/919136899925" style="display:inline-block;background:#E8352A;color:#ffffff;padding:12px 30px;border-radius:25px;text-decoration:none;font-weight:bold;">Chat on WhatsApp</a>
</td></tr>
<tr><td style="background:#0a1a1a;padding:15px;text-align:center;font-size:12px;color:#888;">
&copy; 2026 RoboPirate | WE Smart Lab | <a href="https://robopirate.in" style="color:#0D9B8A;">robopirate.in</a>
</td></tr>
</table>
</td></tr>
</table>
</body>
</html>"""


@dataclass
class Recipient:
    id: int
    sequence_id: str
    email: str
    name: str
    org: str
    extra_json: str
    imported_at: str


@dataclass
class BatchResult:
    queued: int
    sent: int = 0
    drafted: int = 0
    error: Optional[str] = None


class CampaignEngine:
    def __init__(self, db: Database, gmail: GmailClient, ollama_url="http://localhost:11434"):
        self.db = db
        self.gmail = gmail
        self.ollama_url = ollama_url
        self._running = False
        self._thread = None
        self._paused = False
        self._last_batch_process_time = None
        self.brief_email = db.get_meta("brief_email") or ""
        self.calendar = CalendarManager() if CALENDAR_AVAILABLE else None
        self.drive = DriveManager() if DRIVE_AVAILABLE else None
        self.logs = deque(maxlen=200)
        self._log_callbacks = []

    def add_log_callback(self, fn):
        self._log_callbacks.append(fn)

    def _log(self, msg: str):
        ts = datetime.now().strftime("%H:%M:%S")
        line = f"[{ts}] {msg}"
        self.logs.append(line)
        print(f"[Engine] {msg}")
        for fn in self._log_callbacks:
            try:
                fn(line)
            except:
                pass

    # -- Lifecycle --
    def start(self):
        if self._running: return
        self._running = True

        # RESUME-ON-BOOT: Check for batches stuck in "running" status
        try:
            running_batches = self.db.get_running_batches()
            if running_batches:
                self._log(f"[RESUME] Found {len(running_batches)} batch(es) in RUNNING status from previous session")
                for batch in running_batches:
                    self._log(f"[RESUME] Will continue batch '{batch['name']}' (ID: {batch['id']})")
            else:
                self._log("[RESUME] No running batches from previous session")

            # Also check scheduled batches that may have missed their time
            scheduled_batches = self.db.get_scheduled_batches()
            if scheduled_batches:
                now = datetime.now()
                missed = 0
                for batch in scheduled_batches:
                    sched_str = batch.get("scheduled_at")
                    if sched_str:
                        try:
                            sched_dt = datetime.fromisoformat(sched_str)
                            if now > sched_dt:
                                missed += 1
                                self._log(f"[RESUME] Batch '{batch['name']}' missed schedule ({sched_dt.strftime('%d %b %H:%M')}) — will auto-start")
                        except:
                            pass
                if missed > 0:
                    self._log(f"[RESUME] {missed} scheduled batch(es) missed their time while system was off")
        except Exception as e:
            self._log(f"[RESUME] Error checking previous state: {e}")

        self._thread = threading.Thread(target=self._loop, daemon=True)
        self._thread.start()
        self._log("Raj Engine started")

    def stop(self):
        self._running = False
        self._log("Engine stopping...")

    def pause(self):
        self._paused = True
        self._log("PAUSED")

    def resume(self):
        self._paused = False
        self._log("RESUMED")

    def is_running(self): return self._running
    def is_paused(self): return self._paused

    # -- Main Loop --
    def _loop(self):
        while self._running:
            try:
                if not self._paused:
                    self._tick()
            except Exception as e:
                self._log(f"LOOP ERROR: {e}")
            time.sleep(60)

    def _tick(self):
        now = datetime.now()
        self._check_scheduled_sends(now)
        self._process_running_batches(now)
        self._check_auto_start_scheduled_batches(now)
        self._check_bounce_scan(now)
        self._check_reply_scan(now)
        self._check_emergency_commands(now)
        self._check_eod(now)
        self._check_morning_brief(now)

    # -- Batch Processing (NEW) --
    def _process_running_batches(self, now: datetime):
        """Process running batches and send emails at staggered intervals."""
        try:
            running_batches = self.db.execute(
                "SELECT * FROM batches WHERE status='running' ORDER BY created_at"
            ).fetchall()

            if not running_batches:
                return

            # 🔒 BATCH SLOT LOCK: Only process one batch at a time
            if hasattr(self, '_last_batch_process_time') and self._last_batch_process_time:
                if (now - self._last_batch_process_time).total_seconds() < 2:
                    return  # Wait 2 seconds between batch sends
            self._last_batch_process_time = now

            for batch_row in running_batches:
                batch = dict(batch_row)
                batch_id = batch["id"]
                seq_id = batch["sequence_id"]
                stagger = batch.get("stagger_minutes", 0) or 1
                day_offset = batch.get("day_offset", 1)

                # Find next pending recipient
                next_recipient = self.db.execute("""
                    SELECT r.id, r.sequence_id, r.email, r.name, r.org, r.extra_json, r.imported_at
                    FROM recipients r
                    JOIN batch_recipients br ON r.id = br.recipient_id
                    WHERE br.batch_id = ? AND br.status = 'pending'
                    ORDER BY r.id
                    LIMIT 1
                """, (batch_id,)).fetchone()

                if not next_recipient:
                    # All sent - mark completed and auto-advance
                    self.db.batch_update_status(batch_id, "completed")
                    self._log(f"[Batch {batch_id}] Completed: all recipients sent")
                    self._auto_advance_batch(batch)
                    continue

                # Check if enough time passed since last send
                last_send = self.db.execute("""
                    SELECT MAX(sent_at) FROM batch_recipients 
                    WHERE batch_id = ? AND status = 'sent'
                """, (batch_id,)).fetchone()[0]

                if last_send:
                    last_dt = datetime.fromisoformat(last_send)
                    minutes_since = (now - last_dt).total_seconds() / 60
                    if minutes_since < stagger:
                        continue  # Wait longer

                # BLACKLIST CHECK: Skip if email was blacklisted after Day 1
                rec_email = next_recipient[2]  # email is 3rd column in SELECT
                if self.db.blacklist_has(rec_email):
                    self._log(f"[Batch {batch_id}] SKIPPING blacklisted: {rec_email}")
                    self.db.execute("""
                        UPDATE batch_recipients SET status='skipped' 
                        WHERE batch_id=? AND recipient_id=?
                    """, (batch_id, next_recipient[0]))
                    self.db.commit()
                    continue

                # 🚫 SUNDAY FILTER: Skip sends on Sunday
                if now.weekday() == 6:  # Sunday
                    self._log(f"[Batch {batch_id}] SUNDAY — skipping send for {rec_email}, will resume Monday")
                    continue

                # Send email
                rec = Recipient(*next_recipient[:7])
                subj, body = self.render(seq_id, day_offset, rec)
                if not subj:
                    self._log(f"[Batch {batch_id}] No template for {rec.email} Day {day_offset}, skipping")
                    self.db.execute("UPDATE batch_recipients SET status='failed' WHERE batch_id=? AND recipient_id=?",
                                    (batch_id, rec.id))
                    continue

                try:
                    msg = self.gmail.send_email(rec.email, subj, body)
                    self.db.execute("""
                        UPDATE batch_recipients SET status='sent', sent_at=?
                        WHERE batch_id=? AND recipient_id=?
                    """, (now.isoformat(), batch_id, rec.id))
                    self.db.campaign_queue_send(rec.id, day_offset, subj, msg.get("id"), "sent", batch_id)
                    self._log(f"[Batch {batch['name']}] Sent to {rec.email} ({seq_id.upper()} Day {day_offset})")
                except Exception as e:
                    self._log(f"[Batch {batch_id}] Failed to send to {rec.email}: {e}")
                    self.db.execute("UPDATE batch_recipients SET status='failed' WHERE batch_id=? AND recipient_id=?",
                                    (batch_id, rec.id))

        except Exception as e:
            self._log(f"DEBUG ERROR in _process_running_batches: {e}")
            import traceback
            self._log(f"DEBUG TRACEBACK: {traceback.format_exc()}")

    def _auto_advance_batch(self, completed_batch: dict):
        """Auto-create next day batch from POOL when current completes."""
        seq_id = completed_batch["sequence_id"]
        current_day = completed_batch.get("day_offset", 1)
        cfg = SEQUENCES.get(seq_id)
        if not cfg:
            return

        days = cfg["days"]
        try:
            idx = days.index(current_day)
        except ValueError:
            return

        if idx >= len(days) - 1:
            self._log(f"[AUTO-ADVANCE] {seq_id.upper()} sequence complete! All {len(days)} days done.")
            return

        next_day = days[idx + 1]
        parent_name = completed_batch["name"]
        base_name = parent_name.split("-D")[0] if "-D" in parent_name else parent_name
        next_name = f"{base_name}-D{next_day}"

        # Schedule for +2 days at 10 AM
        scheduled = (datetime.now() + timedelta(days=2)).replace(hour=10, minute=0, second=0, microsecond=0)

        # Get parent_batch_id (link to original batch)
        parent_batch_id = completed_batch.get("parent_batch_id") or completed_batch["id"]

        # Get recipients from the completed batch to copy to next batch
        prev_recipients = self.db.batch_get_recipients(completed_batch["id"])

        # Create new batch with parent link
        new_batch_id = self.db.batch_create(
            next_name, seq_id, scheduled.isoformat(),
            stagger_minutes=completed_batch.get("stagger_minutes", 2),
            day_offset=next_day,
            parent_batch_id=parent_batch_id
        )

        # Copy all recipients from parent batch
        for r in prev_recipients:
            self.db.batch_add_recipient(new_batch_id, r["id"])

        self._log(f"[AUTO-ADVANCE] Created {next_name} for {scheduled.strftime('%d %b %H:%M')} ({len(prev_recipients)} recipients)")
        self._log(f"[AUTO-ADVANCE] Pipeline: {base_name} Day {current_day} → Day {next_day} (parent: {parent_batch_id})")

    def _check_auto_start_scheduled_batches(self, now: datetime):
        """Auto-start scheduled batches when their time arrives."""
        scheduled = self.db.execute("""
            SELECT * FROM batches 
            WHERE status='scheduled' AND scheduled_at IS NOT NULL
        """).fetchall()

        for batch_row in scheduled:
            batch = dict(batch_row)
            sched_str = batch.get("scheduled_at")
            if not sched_str:
                continue
            try:
                sched_dt = datetime.fromisoformat(sched_str)
                if now >= sched_dt:
                    self.db.batch_update_status(batch["id"], "running")
                    self._log(f"[AUTO-START] Batch '{batch['name']}' is now running")
            except:
                pass

    # -- Scheduled Sends (10 AM, auto-send sequences) --
    def _check_scheduled_sends(self, now: datetime):
        # 🚫 SUNDAY FILTER: No scheduled sends on Sunday
        if now.weekday() == 6:
            return

        if now.hour != 10 or now.minute > 5:
            return
        today = now.strftime("%Y-%m-%d")
        if self.db.get_meta("last_scheduled_send_date") == today:
            return

        self._log(f"Scheduled send check: {today}")
        for seq_id in SEQUENCES:
            if self.db.get_meta(f"pause_{seq_id}") == "true":
                self._log(f"{seq_id.upper()} is paused, skipping")
                continue
            for day in SEQUENCES[seq_id]["days"]:
                due = self.due_recipients(seq_id, day)
                if due:
                    self._log(f"{seq_id.upper()} Day {day}: {len(due)} due. Auto-sending...")
                    result = self.send_batch(seq_id, day)
                    self._log(f"Sent {result.sent}/{result.queued}")
        self.db.set_meta("last_scheduled_send_date", today)

    # -- Import --
    def smart_import(self, filepath: str, sequence_id: str) -> dict:
        """Smart import to POOL only (no batch creation). Leads go to DB first."""
        if not SMART_IMPORT_AVAILABLE:
            return {"success": False, "error": "smart_importer.py not available"}
        try:
            importer = SmartImporter(self.db, self)
            return importer.import_to_pool(filepath, sequence_id)
        except Exception as e:
            self._log(f"Smart import error: {e}")
            return {"success": False, "error": str(e)}

    def import_recipients(self, path: str, sequence_id: str, mapping: dict) -> Tuple[int, int]:
        try:
            import openpyxl
        except ImportError:
            self._log("openpyxl not installed. Run: pip install openpyxl")
            return 0, 0

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        imported, skipped = 0, 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            email = str(row_dict.get(mapping.get("email", "Email"), "")).strip().lower()
            name = str(row_dict.get(mapping.get("name", "Name"), "")).strip()
            org = str(row_dict.get(mapping.get("org", "Organization"), "")).strip()

            if not email or "@" not in email:
                skipped += 1; continue
            if self.db.blacklist_has(email):
                skipped += 1; continue

            extra = {k: v for k, v in row_dict.items() if k not in mapping.values()}
            try:
                self.db.execute("INSERT INTO recipients (sequence_id, email, name, org, extra_json) VALUES (?, ?, ?, ?, ?)",
                                (sequence_id, email, name, org, json.dumps(extra)))
                imported += 1
            except:
                skipped += 1
        self.db.commit()
        self._log(f"Imported {imported} leads, skipped {skipped}")
        return imported, skipped

    def import_blacklist(self, emails: List[str], reason: str = "imported"):
        count = 0
        for email in emails:
            email = email.strip().lower()
            if email and "@" in email:
                self.db.blacklist_add(email, reason)
                count += 1
        self._log(f"Imported {count} blacklisted emails")
        return count

    # -- Templates --
    def sync_templates(self) -> dict:
        self._log("Syncing templates from Gmail...")
        drafts = self.gmail.list_drafts(100)
        loaded = 0
        found_names = []
        skipped = []

        for d in drafts:
            subject = d.get("subject", "")
            found_names.append(subject)
            draft_id = d.get("id", "")

            m = re.search(r"(SCHOOL|CSR)[\s\-]*EMAIL[\s\-]*(\d+)", subject, re.IGNORECASE)
            if not m:
                m = re.search(r"(SCHOOL|CSR)[\s\-]*(\d+)", subject, re.IGNORECASE)

            if not m:
                skipped.append(f"No match: {subject}")
                continue

            seq = m.group(1).lower()
            num = int(m.group(2))
            day = EMAIL_NUM_TO_DAY.get(num)
            if day is None:
                skipped.append(f"Invalid day num {num}: {subject}")
                continue

            if seq not in SEQUENCES:
                skipped.append(f"Unknown seq {seq}: {subject}")
                continue

            if day not in SEQUENCES[seq]["days"]:
                skipped.append(f"Invalid day {day} for {seq}: {subject}")
                continue

            # RESPECT LOCK STATUS
            if self.is_template_locked(seq, day):
                skipped.append(f"Locked: {seq.upper()} Day {day} - skipping")
                self._log(f"Skipping locked template: {seq.upper()} Day {day}")
                continue

            full = self.gmail.get_draft_full(draft_id)
            if full and full.get("html_body"):
                self.db.template_put(seq, day, full["subject"], full["html_body"])
                loaded += 1
                self._log(f"Loaded: {subject} -> {seq.upper()} Day {day}")
            else:
                skipped.append(f"Failed to fetch body: {subject} (draft_id={draft_id})")
                self._log(f"WARNING: Found matching draft but could not fetch body: {subject}")

        missing = []
        for seq_id, cfg in SEQUENCES.items():
            for day in cfg["days"]:
                if self.db.template_get(seq_id, day) is None:
                    missing.append(f"{seq_id.upper()} Day {day}")

        self._log(f"Sync complete: {loaded} loaded, {len(skipped)} skipped, {len(missing)} missing")
        if skipped:
            for s in skipped[:10]:
                self._log(f"  Skip reason: {s}")

        return {"loaded": loaded, "missing": missing, "found_names": found_names, "skipped": skipped}

    # -- Template Locking System --
    def lock_templates(self) -> dict:
        locked = 0
        for seq_id in SEQUENCES:
            for day in SEQUENCES[seq_id]["days"]:
                tmpl = self.db.template_get(seq_id, day)
                if tmpl:
                    self.db.set_meta(f"locked_{seq_id}_{day}", "true")
                    locked += 1
        self._log(f"Locked {locked} templates. Sync will not overwrite locked templates.")
        return {"locked": locked}

    def unlock_template(self, seq_id: str, day: int):
        self.db.set_meta(f"locked_{seq_id}_{day}", "false")
        self._log(f"Unlocked {seq_id.upper()} Day {day} for updates")

    def is_template_locked(self, seq_id: str, day: int) -> bool:
        return self.db.get_meta(f"locked_{seq_id}_{day}") == "true"

    def create_missing_drafts(self) -> dict:
        created = []
        for seq_id in SEQUENCES:
            for day in SEQUENCES[seq_id]["days"]:
                if self.db.template_get(seq_id, day) is None:
                    tmpl = self.generate_template(seq_id, day)
                    if "error" not in tmpl:
                        self.db.template_put(seq_id, day, tmpl["subject"], tmpl["html_body"])
                        try:
                            draft = self.gmail.draft_email(
                                "om@robopirate.in",
                                f"[TEMPLATE] {tmpl['subject']}",
                                tmpl["html_body"]
                            )
                            created.append(f"{seq_id.upper()} Day {day}")
                            self._log(f"Created draft for {seq_id.upper()} Day {day} — review in Gmail")
                        except Exception as e:
                            self._log(f"DB saved but Gmail draft failed for {seq_id.upper()} Day {day}: {e}")
        return {"created": created, "count": len(created)}

    def get_template_status(self) -> dict:
        status = {}
        for seq_id in SEQUENCES:
            status[seq_id] = {}
            for day in SEQUENCES[seq_id]["days"]:
                tmpl = self.db.template_get(seq_id, day)
                locked = self.is_template_locked(seq_id, day)
                if tmpl:
                    source = self.db.get_meta(f"source_{seq_id}_{day}") or "unknown"
                    status[seq_id][day] = {
                        "exists": True,
                        "locked": locked,
                        "source": source,
                        "subject": tmpl["subject"][:60]
                    }
                else:
                    status[seq_id][day] = {
                        "exists": False,
                        "locked": False,
                        "source": None,
                        "subject": None
                    }
        return status

    def get_templates(self) -> dict:
        out = {}
        for seq_id in SEQUENCES:
            out[seq_id] = {}
            for day in SEQUENCES[seq_id]["days"]:
                t = self.db.template_get(seq_id, day)
                out[seq_id][day] = t
        return out

    # -- Generate Missing Template --
    def generate_template(self, seq_id: str, day: int) -> dict:
        cfg = SEQUENCES.get(seq_id)
        if not cfg:
            return {"error": "Invalid sequence"}

        assets = cfg.get("assets", {}).get(day, {})
        persona = cfg.get("persona", "school")

        content = self._generate_content(seq_id, day, assets)
        subject = self._generate_subject(seq_id, day)

        html = HTML_TEMPLATE.format(title=subject, content=content)

        return {
            "subject": subject,
            "html_body": html,
            "seq_id": seq_id,
            "day": day,
            "assets_used": list(assets.keys())
        }

    def _generate_subject(self, seq_id: str, day: int) -> str:
        subjects = {
            "school": {
                1: "{{SCHOOL_NAME}} — Transform Your School with Hands-On STEM Labs",
                3: "{{SCHOOL_NAME}} — NEP 2020 Compliance: Is Your School Ready?",
                5: "{{PRINCIPAL_NAME}}, See How {{SCHOOL_NAME}} Can Lead STEM Education",
                7: "{{SCHOOL_NAME}} — Join 85+ Schools Already Using WSL",
                10: "{{PRINCIPAL_NAME}}, Final Call: WSL Subscription Plans for {{SCHOOL_NAME}}"
            },
            "csr": {
                1: "{{COMPANY_NAME}} — CSR Impact: 65,000+ Students Reached",
                3: "{{COMPANY_NAME}} — Schedule VII Alignment + STEM Education",
                5: "{{CSR_HEAD_NAME}}, Sangli Success Story for {{COMPANY_NAME}}",
                7: "{{COMPANY_NAME}} — FY Budget Planning: STEM Investment ROI",
                10: "{{CSR_HEAD_NAME}}, Partner with RoboPirate: Company Profile for {{COMPANY_NAME}}"
            }
        }
        return subjects.get(seq_id, {}).get(day, f"RoboPirate {seq_id.upper()} - Day {day}")

    def _generate_content(self, seq_id: str, day: int, assets: dict) -> str:
        if seq_id == "school":
            return self._generate_school_content(day, assets)
        else:
            return self._generate_csr_content(day, assets)

    def _generate_school_content(self, day: int, assets: dict) -> str:
        contents = {
            1: f"""<p>Dear Principal,</p>
<p>Imagine your students building robots, coding drones, and exploring AI — all within your school walls. For the 2026-27 academic year, this is no longer optional.</p>
<p><strong>WE Smart Lab</strong> by RoboPirate brings cutting-edge STEAM/AI education to Indian schools. We're already in <strong style="color:#F5A623;">85+ labs</strong> across <strong style="color:#F5A623;">6 states</strong>, impacting <strong style="color:#F5A623;">65,000+ students</strong>.</p>
<p>Everything is included — lab setup, 120+ DIY kits, full-time trained teacher, NEP 2020 aligned curriculum, LMS portal, and ongoing support. Schools simply open the door; we handle the rest.</p>
<p style="margin:20px 0;"><a href="{assets.get('brochure', '#')}" style="display:inline-block;background:#0D9B8A;color:#ffffff;padding:12px 25px;border-radius:25px;text-decoration:none;font-weight:bold;">📄 Download WSL Brochure</a></p>
<p style="margin:15px 0;"><a href="{assets.get('video_wsl', '#')}" style="display:inline-block;background:#E8352A;color:#ffffff;padding:10px 20px;border-radius:20px;text-decoration:none;">▶ Watch WSL Overview Video</a></p>
<p style="margin:15px 0;"><a href="{assets.get('video_abp', '#')}" style="display:inline-block;background:#1A6B6B;color:#ffffff;padding:10px 20px;border-radius:20px;text-decoration:none;">▶ ABP Majha Media Coverage</a></p>
<p style="margin:15px 0;"><a href="{assets.get('video_ig', '#')}" style="display:inline-block;background:#F5A623;color:#ffffff;padding:10px 20px;border-radius:20px;text-decoration:none;">▶ WSL Instagram Reel</a></p>
<p>Would you be open to a 15-minute call to discuss how WSL can transform your school?</p>
<p>Best regards,<br><strong>RoboPirate Team</strong><br>WSL Initiative</p>""",

            3: f"""<p>Dear Principal,</p>
<p>With NEP 2020 now in full implementation and the 2026-27 academic year approaching, schools across India are racing to comply with experiential learning and coding mandates from Class 6.</p>
<p><strong>The question is:</strong> Will your school lead this change or play catch-up?</p>
<p>WSL provides:</p>
<ul>
<li>Ready-to-deploy STEM labs</li>
<li>NEP-aligned curriculum</li>
<li>Teacher training programs</li>
<li>Progress tracking dashboards</li>
</ul>
<p style="margin:20px 0;"><a href="{assets.get('video_abp', '#')}" style="display:inline-block;background:#E8352A;color:#ffffff;padding:10px 20px;border-radius:20px;text-decoration:none;">▶ ABP Majha Coverage on NEP & STEM</a></p>
<p style="margin:15px 0;"><a href="{assets.get('video_ig', '#')}" style="display:inline-block;background:#F5A623;color:#ffffff;padding:10px 20px;border-radius:20px;text-decoration:none;">▶ WSL Instagram Reel</a></p>
<p>Let's discuss how your school can be NEP-ready this academic year.</p>
<p>Best regards,<br><strong>RoboPirate Team</strong><br>WSL Initiative</p>""",

            5: f"""<p>Dear Principal,</p>
<p>Let me share a story that might resonate with you.</p>
<p><strong>Veer Baji Prabhu Vidyalay</strong> — a school much like yours — partnered with us in 2024-25 through our WE Smart Lab program. Today, their students have built 12+ working robots, participated in state-level competitions, and seen measurable improvement in science engagement.</p>
<p>See the full impact data, photos, and videos from their lab in the links below. Every result is documented and verified.</p>
<p style="margin:20px 0;"><a href="{assets.get('report_vbv', '#')}" style="display:inline-block;background:#0D9B8A;color:#ffffff;padding:12px 25px;border-radius:25px;text-decoration:none;font-weight:bold;">📄 Read VBV Student Progress Report</a></p>
<p style="margin:15px 0;"><a href="{assets.get('video_star', '#')}" style="display:inline-block;background:#E8352A;color:#ffffff;padding:10px 20px;border-radius:20px;text-decoration:none;">▶ Star News Marathi Coverage</a></p>
<p style="margin:15px 0;"><a href="{assets.get('folder_vbv', '#')}" style="display:inline-block;background:#1A6B6B;color:#ffffff;padding:10px 20px;border-radius:20px;text-decoration:none;">📁 View VBV Event Photos & Videos</a></p>
<p style="margin:15px 0;"><a href="{assets.get('video_ig', '#')}" style="display:inline-block;background:#F5A623;color:#ffffff;padding:10px 20px;border-radius:20px;text-decoration:none;">▶ WSL Instagram Reel</a></p>
<p>Your school could be our next success story.</p>
<p>Best regards,<br><strong>RoboPirate Team</strong><br>WSL Initiative</p>""",

            7: f"""<p>Dear Principal,</p>
<p>You're not alone in this journey. <strong>85+ schools</strong> across Maharashtra, Karnataka, Gujarat, and more have already chosen WSL.</p>
<p>Here's what principals are saying:</p>
<blockquote style="border-left:4px solid #0D9B8A;padding-left:15px;color:#555;">
"WSL transformed how our students engage with science. The hands-on approach is exactly what NEP envisioned."<br>
-- Principal, Pune District
</blockquote>
<p style="margin:20px 0;"><a href="{assets.get('profile', '#')}" style="display:inline-block;background:#0D9B8A;color:#ffffff;padding:12px 25px;border-radius:25px;text-decoration:none;font-weight:bold;">📄 Download RoboPirate Company Profile</a></p>
<p style="margin:15px 0;"><a href="{assets.get('video_abp', '#')}" style="display:inline-block;background:#E8352A;color:#ffffff;padding:10px 20px;border-radius:20px;text-decoration:none;">▶ ABP Majha Coverage</a></p>
<p style="margin:15px 0;"><a href="{assets.get('video_star', '#')}" style="display:inline-block;background:#1A6B6B;color:#ffffff;padding:10px 20px;border-radius:20px;text-decoration:none;">▶ Star News Marathi</a></p>
<p style="margin:15px 0;"><a href="{assets.get('video_ig', '#')}" style="display:inline-block;background:#F5A623;color:#ffffff;padding:10px 20px;border-radius:20px;text-decoration:none;">▶ WSL Instagram Reel</a></p>
<p>Ready to join them?</p>
<p>Best regards,<br><strong>RoboPirate Team</strong><br>WSL Initiative</p>""",

            10: f"""<p>Dear Principal,</p>
<p>This is my final email for the 2026-27 academic year planning. With admissions season approaching, I don't want your students to miss this opportunity.</p>
<p>We've prepared flexible WE Smart Lab subscription plans for schools of all sizes:</p>
<ul>
<li><strong>WSL20:</strong> 400 students | 1 trainer | 1 lab | ~Rs.8.2L/year</li>
<li><strong>WSL30:</strong> 600 students | 1 trainer | 1 lab | ~Rs.10.2L/year</li>
<li><strong>WSL30 V.2:</strong> 800 students | 2 trainers | 1 lab | ~Rs.12.8L/year</li>
</ul>
<p>Every plan includes: complete lab setup, 120+ DIY kits, full-time trained teacher, NEP 2020 + NCF aligned curriculum, LMS portal, assessments, and ongoing support.</p>
<p><a href="{assets.get('plans', '#')}" style="color:#0D9B8A;">📄 View WE Smart Lab Subscription Plans</a></p>
<p>Or explore everything at <a href="https://robopirate.in" style="color:#0D9B8A;">robopirate.in</a></p>
<p>If now isn't the right time, I understand. But if you're even slightly curious, let's have a 10-minute conversation. No obligation.</p>
<p>Click the WhatsApp button below or reply to this email.</p>
<p>Best regards,<br><strong>RoboPirate Team</strong><br>WSL Initiative</p>"""
        }
        return contents.get(day, f"<p>Template content for Day {day}</p>")

    def _generate_csr_content(self, day: int, assets: dict) -> str:
        contents = {
            1: f"""<p>Dear CSR Head,</p>
<p>Your CSR budget has the power to change <strong>thousands</strong> of young lives.</p>
<p>RoboPirate's <strong>WE Smart Lab</strong> sets up fully managed STEAM/AI Smart Labs inside schools across India. As of May 2026, we've reached <strong style="color:#F5A623;">65,000+ students</strong> across <strong style="color:#F5A623;">6 states</strong> with <strong style="color:#F5A623;">85+ labs</strong> delivered through strategic CSR partnerships.</p>
<p>Each lab includes complete setup, 120+ DIY kits (robotics, coding, AI, ML, IoT, 3D printing), a full-time trained on-site teacher, NEP 2020 + NCF aligned curriculum, LMS portal with weekly assessments, and ongoing technical support.</p>
<p style="margin:20px 0;"><a href="{assets.get('report_sangli1', '#')}" style="display:inline-block;background:#0D9B8A;color:#ffffff;padding:12px 25px;border-radius:25px;text-decoration:none;font-weight:bold;">📄 Sangli Impact Report (Phase 1)</a></p>
<p style="margin:15px 0;"><a href="{assets.get('video_abp', '#')}" style="display:inline-block;background:#E8352A;color:#ffffff;padding:10px 20px;border-radius:20px;text-decoration:none;">▶ ABP Majha Coverage</a></p>
<p style="margin:15px 0;"><a href="{assets.get('video_sangli', '#')}" style="display:inline-block;background:#1A6B6B;color:#ffffff;padding:10px 20px;border-radius:20px;text-decoration:none;">▶ Sangli 15-Day Workshop Video</a></p>
<p style="margin:15px 0;"><a href="{assets.get('video_ig', '#')}" style="display:inline-block;background:#F5A623;color:#ffffff;padding:10px 20px;border-radius:20px;text-decoration:none;">▶ WSL Instagram Reel</a></p>
<p>Would you be open to exploring how your CSR mandate can create measurable STEM impact?</p>
<p>Best regards,<br><strong>RoboPirate CSR Team</strong></p>""",

            3: f"""<p>Dear CSR Head,</p>
<p>Schedule VII of the Companies Act explicitly supports:</p>
<ul>
<li>Education (item ii)</li>
<li>Skill development (item x)</li>
<li>Rural development (item xii)</li>
</ul>
<p>WSL aligns perfectly with all three.</p>
<p style="margin:20px 0;"><a href="{assets.get('report_sangli1', '#')}" style="display:inline-block;background:#0D9B8A;color:#ffffff;padding:12px 25px;border-radius:25px;text-decoration:none;font-weight:bold;">📄 Sangli Impact Report</a></p>
<p style="margin:15px 0;"><a href="{assets.get('brochure', '#')}" style="display:inline-block;background:#1A6B6B;color:#ffffff;padding:10px 20px;border-radius:20px;text-decoration:none;">📄 WSL Bifold Brochure</a></p>
<p style="margin:15px 0;"><a href="{assets.get('video_ig', '#')}" style="display:inline-block;background:#F5A623;color:#ffffff;padding:10px 20px;border-radius:20px;text-decoration:none;">▶ WSL Instagram Reel</a></p>
<p>Our programs are designed for maximum social return on investment (SROI) with full compliance documentation.</p>
<p>Best regards,<br><strong>RoboPirate CSR Team</strong></p>""",

            5: f"""<p>Dear CSR Head,</p>
<p>Numbers tell stories better than words.</p>
<p><strong>Sangli District Phase 2 Results — WE Smart Lab Impact (Delivered 2025-26):</strong></p>
<ul>
<li>15 schools equipped with fully managed STEAM/AI labs</li>
<li>4,500+ students trained in robotics, coding, AI & IoT</li>
<li>87% teacher satisfaction rate</li>
<li>3 students won state-level competitions</li>
<li>1.5L+ student projects completed across all programs</li>
</ul>
<p>All data is from real, on-ground delivery. Reports, photos, and videos are available below.</p>
<p style="margin:20px 0;"><a href="{assets.get('report_sangli2', '#')}" style="display:inline-block;background:#0D9B8A;color:#ffffff;padding:12px 25px;border-radius:25px;text-decoration:none;font-weight:bold;">📄 Sangli Report Phase 2</a></p>
<p style="margin:15px 0;"><a href="{assets.get('report_vbv', '#')}" style="display:inline-block;background:#1A6B6B;color:#ffffff;padding:10px 20px;border-radius:20px;text-decoration:none;">📄 VBV Student Progress Report</a></p>
<p style="margin:15px 0;"><a href="{assets.get('video_star', '#')}" style="display:inline-block;background:#E8352A;color:#ffffff;padding:10px 20px;border-radius:20px;text-decoration:none;">▶ Star News Marathi Coverage</a></p>
<p style="margin:15px 0;"><a href="{assets.get('folder_sangli', '#')}" style="display:inline-block;background:#1A6B6B;color:#ffffff;padding:10px 20px;border-radius:20px;text-decoration:none;">📁 Sangli Event Photos & Videos</a></p>
<p style="margin:15px 0;"><a href="{assets.get('video_ig', '#')}" style="display:inline-block;background:#F5A623;color:#ffffff;padding:10px 20px;border-radius:20px;text-decoration:none;">▶ WSL Instagram Reel</a></p>
<p>This could be your company's legacy.</p>
<p>Best regards,<br><strong>RoboPirate CSR Team</strong></p>""",

            7: f"""<p>Dear CSR Head,</p>
<p>FY 2026-27 budget season is here — May 2026 is when CSR allocations are locked. Where will your CSR rupees create the most impact?</p>
<p>Consider the WE Smart Lab model:</p>
<ul>
<li>Setup cost: Rs.2.5L – 8L per school (one-time, based on tier)</li>
<li>Annual program cost: Rs.7L per school (CSR School Model)</li>
<li>Cost per student impacted: Under Rs.500/year</li>
<li>Tax benefits: 100% deductible under Companies Act 2013 Schedule VII</li>
<li>Full compliance documentation + quarterly impact reports included</li>
</ul>
<p>View detailed plans at <a href="https://robopirate.in" style="color:#0D9B8A;">robopirate.in</a> or download the subscription PDF below.</p>
<p style="margin:20px 0;"><a href="{assets.get('plans', '#')}" style="display:inline-block;background:#0D9B8A;color:#ffffff;padding:12px 25px;border-radius:25px;text-decoration:none;font-weight:bold;">📄 WE Smart Lab Subscription Plans</a></p>
<p style="margin:15px 0;"><a href="{assets.get('video_wsl', '#')}" style="display:inline-block;background:#E8352A;color:#ffffff;padding:10px 20px;border-radius:20px;text-decoration:none;">▶ WSL Overview Video</a></p>
<p style="margin:15px 0;"><a href="{assets.get('video_abp', '#')}" style="display:inline-block;background:#1A6B6B;color:#ffffff;padding:10px 20px;border-radius:20px;text-decoration:none;">▶ ABP Majha Coverage</a></p>
<p style="margin:15px 0;"><a href="{assets.get('video_sangli', '#')}" style="display:inline-block;background:#1A6B6B;color:#ffffff;padding:10px 20px;border-radius:20px;text-decoration:none;">▶ Sangli 15-Day Workshop</a></p>
<p style="margin:15px 0;"><a href="{assets.get('video_ig', '#')}" style="display:inline-block;background:#F5A623;color:#ffffff;padding:10px 20px;border-radius:20px;text-decoration:none;">▶ WSL Instagram Reel</a></p>
<p>Let's discuss a pilot program for Q1.</p>
<p>Best regards,<br><strong>RoboPirate CSR Team</strong></p>""",

            10: f"""<p>Dear CSR Head,</p>
<p>This is my final outreach for FY 2026-27 planning. With budgets being locked in May 2026, I respect your time and decision.</p>
<p>If you've been considering STEM education as part of your CSR strategy, let's not let another quarter pass.</p>
<p>Download our full credentials:</p>
<p><a href="{assets.get('profile', '#')}" style="color:#0D9B8A;">📄 RoboPirate Company Profile</a></p>
<p><a href="{assets.get('kits', '#')}" style="color:#0D9B8A;">📦 Kits In The Box — Product Catalog</a></p>
<p>Learn more at <a href="https://robopirate.in" style="color:#0D9B8A;">robopirate.in</a> or message us directly on WhatsApp below.</p>
<p>I'm available for a 20-minute presentation at your office or via video call. No pitch, just facts and possibilities.</p>
<p>Click the WhatsApp button below or simply reply "Interested" and I'll send available slots.</p>
<p>Best regards,<br><strong>RoboPirate CSR Team</strong></p>"""
        }
        return contents.get(day, f"<p>Template content for Day {day}</p>")

    def save_generated_template(self, seq_id: str, day: int) -> bool:
        template = self.generate_template(seq_id, day)
        if "error" in template:
            self._log(f"Failed to generate {seq_id.upper()} Day {day}: {template['error']}")
            return False

        self.db.template_put(seq_id, day, template["subject"], template["html_body"], "generated")

        try:
            draft = self.gmail.draft_email(
                "om@robopirate.in",
                f"[TEMPLATE] {template['subject']}",
                template["html_body"]
            )
            self._log(f"Generated {seq_id.upper()} Day {day} template + Gmail draft created")
            return True
        except Exception as e:
            self._log(f"Saved to DB but Gmail draft failed: {e}")
            return True

    # -- Due Recipients --
    def due_recipients(self, sequence_id: str, day: int, limit=None) -> List[Recipient]:
        cfg = SEQUENCES.get(sequence_id)
        if not cfg or day not in cfg["days"]: return []

        idx = cfg["days"].index(day)
        if idx == 0:
            sql = """SELECT r.* FROM recipients r WHERE r.sequence_id=? 
                     AND NOT EXISTS (SELECT 1 FROM sends s WHERE s.recipient_id=r.id)
                     AND NOT EXISTS (SELECT 1 FROM blacklist b WHERE b.email=r.email)
                     ORDER BY r.id"""
            params = (sequence_id,)
        else:
            prev = cfg["days"][idx - 1]
            gap = day - prev
            cutoff = (datetime.now() - timedelta(days=gap)).isoformat()
            sql = """SELECT r.* FROM recipients r
                     JOIN sends s ON s.recipient_id=r.id AND s.day=? AND s.status IN ('sent','drafted')
                     WHERE r.sequence_id=? AND s.created_at<=?
                     AND NOT EXISTS (SELECT 1 FROM sends s2 WHERE s2.recipient_id=r.id AND s2.day=?)
                     AND NOT EXISTS (SELECT 1 FROM blacklist b WHERE b.email=r.email)
                     AND NOT EXISTS (SELECT 1 FROM sends s3 WHERE s3.recipient_id=r.id AND s3.status='replied')
                     ORDER BY s.created_at"""
            params = (prev, sequence_id, cutoff, day)

        rows = self.db.execute(sql, params).fetchall()
        return [Recipient(*r) for r in rows][:limit] if limit else [Recipient(*r) for r in rows]

    # -- Render --
    def render(self, seq_id: str, day: int, rec: Recipient) -> Tuple[Optional[str], Optional[str]]:
        tmpl = self.db.template_get(seq_id, day)
        if not tmpl: return None, None

        subj, body = tmpl["subject"] or "", tmpl["html_body"] or ""
        extra = json.loads(rec.extra_json or "{}")

        placeholders = {
            "{{PRINCIPAL_NAME}}": rec.name, "{{SCHOOL_NAME}}": rec.org,
            "{{CSR_HEAD_NAME}}": rec.name, "{{COMPANY_NAME}}": rec.org,
            "{{OPENING_LINE}}": extra.get("Opening Line", extra.get("opening_line", "")),
            "{{NAME}}": rec.name, "{{ORG}}": rec.org, "{{EMAIL}}": rec.email,
        }
        for ph, val in placeholders.items():
            subj = subj.replace(ph, str(val))
            body = body.replace(ph, str(val))
        return subj, body

    # -- Send Batch (AUTO-SEND for sequences) --
    def send_batch(self, seq_id: str, day: int, limit=None, dry_run=False) -> BatchResult:
        due = self.due_recipients(seq_id, day, limit)
        if not due: return BatchResult(queued=0, sent=0)
        if dry_run: return BatchResult(queued=len(due), sent=0)

        sent = 0
        for i, rec in enumerate(due):
            subj, body = self.render(seq_id, day, rec)
            if not subj:
                self._log(f"No template for {rec.email}, skipping")
                continue
            try:
                msg = self.gmail.send_email(rec.email, subj, body)
                self.db.campaign_queue_send(rec.id, day, subj, msg.get("id"), "sent")
                sent += 1
                self._log(f"Sent to {rec.email}")
                time.sleep(SEND_DELAY)
            except Exception as e:
                err = str(e)
                if "quota" in err.lower() or "rate" in err.lower() or "limit" in err.lower():
                    self._log("Rate limit hit. Saving remaining to pending_resumes...")
                    for r in due[i:]:
                        rs, rb = self.render(seq_id, day, r)
                        self.db.execute(
                            "INSERT INTO pending_resumes (sequence_id, day, recipient_id, subject, status, error) VALUES (?, ?, ?, ?, ?, ?)",
                            (seq_id, day, r.id, rs or subj, "pending", err[:200])
                        )
                    self.db.commit()
                    remaining = len(due) - i
                    self._log(f"Saved {remaining} emails to pending_resumes. Type 'resume batch {seq_id} day {day}' to continue.")
                    return BatchResult(queued=len(due), sent=sent, error="quota_hit")
                self._log(f"Failed: {rec.email} -- {e}")
        return BatchResult(queued=len(due), sent=sent)

    # -- Test Send --
    def test_send(self, email: str, seq_id: str, day: int) -> bool:
        tmpl = self.db.template_get(seq_id, day)
        if not tmpl:
            self._log("No template found")
            return False
        try:
            self.gmail.send_email(email, f"[TEST] {tmpl['subject']}", tmpl["html_body"])
            self._log(f"Test sent to {email}")
            return True
        except Exception as e:
            self._log(f"Test failed: {e}")
            return False

    # -- Summary --
    def get_summary(self) -> dict:
        return self.db.get_dashboard_summary()

    def get_catch_up(self) -> List[dict]:
        catch = []
        for seq_id in SEQUENCES:
            for day in SEQUENCES[seq_id]["days"]:
                due = self.due_recipients(seq_id, day)
                if due:
                    overdue = 0
                    if day != 1:
                        prev = SEQUENCES[seq_id]["days"][SEQUENCES[seq_id]["days"].index(day) - 1]
                        oldest = self.db.execute("SELECT MIN(created_at) FROM sends s JOIN recipients r ON r.id=s.recipient_id WHERE r.sequence_id=? AND s.day=?", (seq_id, prev)).fetchone()[0]
                        if oldest:
                            expected = datetime.fromisoformat(oldest) + timedelta(days=(day - prev))
                            overdue = max(0, (datetime.now() - expected).days)
                    catch.append({"sequence": seq_id, "day": day, "count": len(due), "overdue_by_days": overdue})
        return catch


    # -- Batch Pipeline --
    def get_batch_pipeline(self, batch_id: int) -> dict:
        """Get full pipeline for a batch (entire sequence journey)."""
        return self.db.batch_get_pipeline(batch_id)

    def get_all_batch_pipelines(self, sequence_id: str = None) -> list:
        """Get all batch pipelines grouped by original batch."""
        return self.db.batch_get_all_pipelines(sequence_id)

    # -- POOL METHODS (NEW) --
    def get_pool(self, sequence_id: str, limit: int = None) -> list:
        """Get unbatched leads from the pool."""
        return self.db.get_pool(sequence_id, limit)

    def get_pool_count(self, sequence_id: str) -> int:
        """Count unbatched leads in pool."""
        return self.db.get_pool_count(sequence_id)

    def create_batch_from_pool(self, name: str, sequence_id: str, batch_size: int,
                                day_offset: int = 1, scheduled_at: str = None,
                                timezone: str = 'Asia/Kolkata', send_rate: int = 0,
                                stagger_minutes: int = 2) -> dict:
        """Create a batch from unbatched leads in the pool."""
        pool_count = self.get_pool_count(sequence_id)
        if pool_count == 0:
            return {"success": False, "error": f"No unbatched leads in {sequence_id.upper()} pool"}

        batch_id, error = self.db.batch_from_pool(
            name=name,
            sequence_id=sequence_id,
            batch_size=batch_size,
            day_offset=day_offset,
            scheduled_at=scheduled_at,
            timezone=timezone,
            send_rate=send_rate,
            stagger_minutes=stagger_minutes
        )

        if error:
            return {"success": False, "error": error}

        batch = self.db.batch_get(batch_id)
        actual_size = self.db.batch_count_recipients(batch_id)
        self._log(f"[POOL] Created batch '{name}' with {actual_size}/{batch_size} leads from {sequence_id.upper()} pool ({pool_count} available)")
        return {
            "success": True,
            "batch_id": batch_id,
            "name": name,
            "sequence_id": sequence_id,
            "size": actual_size,
            "requested_size": batch_size,
            "pool_remaining": pool_count - actual_size,
            "day_offset": day_offset,
            "scheduled_at": scheduled_at
        }

    # -- Blacklist --
    def blacklist_add(self, email: str, reason: str = "manual"):
        self.db.blacklist_add(email, reason)
        self._log(f"Blacklisted: {email}")

    def blacklist_remove(self, email: str):
        self.db.blacklist_remove(email)
        self._log(f"Removed from blacklist: {email}")

    # -- Bounce Scan --
    def _check_bounce_scan(self, now: datetime):
        last = self.db.get_meta("last_bounce_scan")
        if last and (now - datetime.fromisoformat(last)) < timedelta(hours=BOUNCE_INTERVAL): return
        self.scan_bounces(days_back=15)

    def scan_bounces(self, days_back: int = 1) -> dict:
        """Scan for bounces and auto-replies. Deletes processed emails from Gmail."""
        last = self.db.get_meta("last_bounce_scan")
        if last:
            last_dt = datetime.fromisoformat(last)
            scan_since = max(last_dt, datetime.now() - timedelta(days=days_back))
        else:
            scan_since = datetime.now() - timedelta(days=days_back)

        after_str = scan_since.strftime("%Y/%m/%d")

        # TIGHTENED queries - only search for actual bounces
        queries = [
            f"after:{after_str} (from:mailer-daemon OR from:postmaster OR from:Mail Delivery Subsystem OR from:MAILER-DAEMON)",
            f"after:{after_str} (subject:undelivered OR subject:bounce OR subject:'delivery status' OR subject:'delivery failure' OR subject:'failed delivery' OR subject:'address not found' OR subject:'recipient not found' OR subject:'Mail delivery failed' OR subject:'Returned mail')",
            f"after:{after_str} (subject:out of office OR subject:vacation OR subject:'auto reply' OR subject:'automated response' OR subject:'automatic reply' OR subject:'away from office')",
        ]

        all_msgs = []
        seen_ids = set()

        for q in queries:
            try:
                msgs = self.gmail.search_messages(q, 100)
                for m in msgs:
                    if m['id'] not in seen_ids:
                        seen_ids.add(m['id'])
                        all_msgs.append(m)
            except Exception as e:
                self._log(f"Bounce query failed: {e}")

        self._log(f"Bounce scan: {len(all_msgs)} messages to check")

        new_blacklisted = 0
        auto_reply_count = 0
        protected_count = 0
        deleted_count = 0
        skipped = 0
        processed_this_scan = set()

        for msg in all_msgs:
            subject = msg.get("subject", "").lower()
            body = msg.get("body", "") or ""
            from_addr = msg.get("from", "").lower()
            snippet = msg.get("snippet", "") or ""
            msg_id = msg["id"]

            # Skip our own emails
            if "robopirate" in from_addr and "mailer-daemon" not in from_addr:
                continue

            # Check if bounce or auto-reply
            is_bounce = self._looks_like_bounce(from_addr, subject, body)
            is_auto_reply = self._is_auto_reply(subject, body)

            if is_auto_reply and not is_bounce:
                auto_reply_count += 1
                self._log(f"[AUTO-REPLY] {from_addr[:40]}: {subject[:50]}")
                self._delete_bounce_email(msg_id)
                continue

            if not is_bounce:
                self._delete_bounce_email(msg_id)
                continue

            # Extract bounced emails
            addrs = self._extract_bounced(body) or []

            # Try full message
            try:
                full = self.gmail.get_message_full(msg_id)
                if full:
                    full_addrs = self._extract_bounced(full.get("body", "") or "")
                    for a in full_addrs:
                        if a not in addrs:
                            addrs.append(a)
            except:
                pass

            # Try snippet
            snippet_addrs = self._extract_bounced(snippet)
            for a in snippet_addrs:
                if a not in addrs:
                    addrs.append(a)

            if not addrs:
                self._log(f"[BOUNCE] No address extracted: {subject[:60]}")
                self._delete_bounce_email(msg_id)
                continue

            for addr in addrs:
                addr = addr.lower().strip()

                # Skip invalid/URLs
                if not addr or "@" not in addr:
                    continue
                if addr.endswith((".png", ".jpg", ".jpeg", ".gif", ".svg", ".ico", ".css", ".js")):
                    continue
                if "/" in addr or "?" in addr or "&" in addr:
                    continue
                if addr.startswith(("wght@", "size@", "color@", "font@")):
                    continue
                if addr.endswith("@robopirate.in"):
                    protected_count += 1
                    continue
                if self.db.blacklist_has(addr):
                    if addr not in processed_this_scan:
                        skipped += 1
                        processed_this_scan.add(addr)
                    continue
                if addr in processed_this_scan:
                    continue

                processed_this_scan.add(addr)
                self.db.blacklist_add(addr, "bounce")
                new_blacklisted += 1
                self._log(f"[BLACKLIST] {addr}")

            self._delete_bounce_email(msg_id)
            deleted_count += 1

        self.db.set_meta("last_bounce_scan", datetime.now().isoformat())
        self._log(f"Bounce scan: {new_blacklisted} new blacklisted, {auto_reply_count} auto-replies, {protected_count} protected, {deleted_count} deleted, {skipped} already blacklisted")
        return {
            "new_blacklisted": new_blacklisted,
            "auto_replies": auto_reply_count,
            "protected": protected_count,
            "deleted": deleted_count,
            "skipped": skipped
        }

    def _delete_bounce_email(self, msg_id: str):
        """Delete a bounce email from Gmail. Tries trash first, then delete permanently."""
        try:
            self.gmail.trash_message(msg_id)
        except Exception as e:
            try:
                self.gmail.delete_message(msg_id)
            except:
                self._log(f"Could not delete bounce email {msg_id}: {e}")


    def deep_bounce_scan(self, days: int = 30) -> dict:
        """Deep scan inbox for bounce emails over last N days. Only blacklists TRUE bounces."""
        results = {'found': 0, 'blacklisted': 0, 'protected': 0, 'details': []}
        try:
            after_date = (datetime.now() - timedelta(days=days)).strftime("%Y/%m/%d")
            query = f"after:{after_date} (from:mailer-daemon OR from:postmaster OR 'delivery status notification' OR 'undeliverable' OR 'message not delivered')"
            messages = self.gmail.search_messages(query, max_results=200)
            if not messages:
                self._log(f"[DEEP BOUNCE SCAN] No bounce emails found in last {days} days")
                return results

            self._log(f"[DEEP BOUNCE SCAN] Checking {len(messages)} potential bounce emails (last {days} days)...")

            sent_rows = self.db.execute("SELECT DISTINCT email FROM recipients").fetchall()
            our_emails = {r[0].lower().strip() for r in sent_rows}

            for msg in messages:
                try:
                    from_addr = msg.get('from', '').lower()
                    subject = msg.get('subject', '').lower()
                    body = msg.get('body', '').lower()

                    is_mailer = any(x in from_addr for x in ['mailer-daemon', 'postmaster', 'mail delivery subsystem'])
                    is_bounce_subject = any(x in subject for x in [
                        'delivery status notification', 'undeliverable', 'permanent failure',
                        'message not delivered', 'failure notice', 'returned mail'
                    ])
                    if not (is_mailer or is_bounce_subject):
                        continue

                    results['found'] += 1

                    # Extract bounced email
                    bounced_email = None
                    patterns = [
                        r'final-recipient:\s*rfc822;\s*([^\s<>]+)',
                        r'original-recipient:\s*rfc822;\s*([^\s<>]+)',
                        r'\bto:\s*([^\s<>]+@[^\s<>]+)',
                        r'does not exist[:\s]+([^\s<>]+@[^\s<>]+)',
                    ]
                    for pat in patterns:
                        m = re.search(pat, body)
                        if m:
                            bounced_email = m.group(1).strip()
                            break

                    if not bounced_email:
                        emails_in_body = re.findall(r'[\w.-]+@[\w.-]+\.\w+', body)
                        for e in emails_in_body:
                            if e.lower() in our_emails:
                                bounced_email = e
                                break

                    if not bounced_email:
                        continue

                    bounced_email = bounced_email.lower().strip()
                    bounced_email = re.sub(r'[.,;:!?")\'\'\]]+$', '', bounced_email)

                    if not re.match(r'^[\w.-]+@[\w.-]+\.\w+$', bounced_email):
                        continue
                    if self.db.blacklist_has(bounced_email):
                        continue
                    if bounced_email.endswith('@robopirate.in') or bounced_email == 'itsomkarsinghhh@gmail.com':
                        results['protected'] += 1
                        continue
                    if bounced_email not in our_emails:
                        continue

                    self.db.blacklist_add(bounced_email, f"bounce (deep scan {days}d)")
                    results['blacklisted'] += 1
                    results['details'].append({'email': bounced_email, 'action': 'BLACKLISTED'})
                    self.gmail.trash_message(msg['id'])

                except Exception as e:
                    continue

            self._log(f"[DEEP BOUNCE SCAN] Complete: {results['found']} found, {results['blacklisted']} blacklisted, {results['protected']} protected")
            return results

        except Exception as e:
            self._log(f"[Engine] Deep bounce scan error: {e}")
            return results

    def _looks_like_bounce(self, from_addr: str, subject: str, body: str) -> bool:
        """Quick heuristic check if an email looks like a bounce or auto-reply."""
        from_lower = from_addr.lower()
        subj_lower = subject.lower()
        body_lower = body.lower()

        # Known bounce senders
        bounce_senders = [
            "mailer-daemon", "postmaster", "mail delivery subsystem",
            "daemon", "bounce", "undeliverable", "noreply"
        ]
        for sender in bounce_senders:
            if sender in from_lower:
                return True

        # Bounce subject patterns
        bounce_subjects = [
            "undelivered", "bounce", "delivery status", "delivery failure",
            "failed delivery", "address not found", "recipient not found",
            "returned mail", "mail delivery failed", "message not delivered"
        ]
        for pattern in bounce_subjects:
            if pattern in subj_lower:
                return True

        # Auto-reply subject patterns
        auto_subjects = [
            "out of office", "auto reply", "automated response", "automatic reply",
            "vacation", "on leave", "away from office", "abwesenheitsnotiz"
        ]
        for pattern in auto_subjects:
            if pattern in subj_lower:
                return True

        # Body patterns for bounces
        body_bounce_patterns = [
            "final-recipient", "diagnostic-code", "action: failed",
            "status:", "remote server", "smtp error", "550 ", "551 ",
            "552 ", "553 ", "554 ", "recipient address rejected",
            "user unknown", "no such user", "mailbox unavailable"
        ]
        for pattern in body_bounce_patterns:
            if pattern in body_lower:
                return True

        # Body patterns for auto-replies
        auto_body_patterns = [
            "auto-submitted:", "x-autoreply:", "precedence: auto_reply",
            "x-auto-response-suppress:", "i am currently out of",
            "i will be out of", "i am away", "on vacation until",
            "return on", "back on", "this is an automated"
        ]
        for pattern in auto_body_patterns:
            if pattern in body_lower:
                return True

        return False

    @staticmethod
    def _is_auto_reply(subject: str, body: str) -> bool:
        """Detect if message is an auto-reply/out-of-office/vacation response."""
        subject_lower = subject.lower()
        body_lower = body.lower()

        auto_reply_keywords = [
            "out of office", "out of the office", "away from office", "on vacation",
            "on leave", "auto reply", "automated response", "automatic reply",
            "automatic response", "auto-response", "out of office reply",
            "abwesenheitsnotiz", "risposta automatica", "respuesta automatica",
            "reponse automatique", "automatikus valasz", "automatski odgovor",
            "automatisch antwoord", "automaattinen vastaus", "automatsvar",
            "i am currently out of", "i will be out of", "i am away",
            "not in office", "not available", "currently unavailable",
            "thank you for your email", "we have received your email",
            "this is an automated", "this email is automatically",
            "do not reply to this", "noreply", "no reply",
            "i am on holiday", "i am on vacation", "annual leave",
            "maternity leave", "paternity leave", "sick leave",
            "traveling until", "back on", "return on", "will return",
            "limited access to email", "intermittent access",
            "email access limited", "delayed response"
        ]

        for keyword in auto_reply_keywords:
            if keyword in subject_lower or keyword in body_lower:
                return True

        header_patterns = [
            "auto-submitted:", "x-autoreply:", "x-auto-response-suppress:",
            "precedence: auto_reply", "precedence: bulk",
            "x-mailer: autoreply", "x-autoresponder:",
            "vacation:", "x-vacation:", "autoreply:"
        ]
        for pattern in header_patterns:
            if pattern in body_lower:
                return True

        return False

    def _extract_original_sender(self, subject: str, body: str, msg_or_full: dict = None) -> Optional[str]:
        """Extract the original sender email from an auto-reply or bounce message."""
        patterns = [
            r"Original-From:\s*<?([\w.+-]+@[\w.-]+)>?",
            r"From:\s*<?([\w.+-]+@[\w.-]+)>?",
            r"Sender:\s*<?([\w.+-]+@[\w.-]+)>?",
            r"Reply-To:\s*<?([\w.+-]+@[\w.-]+)>?",
            r"was sent by\s*<?([\w.+-]+@[\w.-]+)>?",
            r"sent by\s*<?([\w.+-]+@[\w.-]+)>?",
            r"original message was sent by\s*<?([\w.+-]+@[\w.-]+)>?",
            r"your message to\s*<?([\w.+-]+@[\w.-]+)>?",
            r"email sent to\s*<?([\w.+-]+@[\w.-]+)>?",
            r"message to\s*<?([\w.+-]+@[\w.-]+)>?\s*was",
        ]

        # Try body first, then full message body, then snippet
        texts = [body]
        if msg_or_full:
            texts.append(msg_or_full.get("body", ""))
            texts.append(msg_or_full.get("snippet", ""))

        for text in texts:
            if not text:
                continue
            for pattern in patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    email = match.group(1).strip().lower()
                    if "@" in email and "mailer-daemon" not in email and "postmaster" not in email:
                        return email
            # Try To: field
            to_match = re.search(r"To:\s*<?([\w.+-]+@[\w.-]+)>?", text, re.IGNORECASE)
            if to_match:
                email = to_match.group(1).strip().lower()
                if "@" in email and "mailer-daemon" not in email and "postmaster" not in email:
                    return email

        return None

    @staticmethod
    def _extract_bounced(text: str) -> List[str]:
        if not text: return []
        addrs = []

        # Pattern 1: Gmail DSN format: Final-Recipient: rfc822; email@domain.com (with space)
        for m in re.finditer(r"Final-Recipient:\s*rfc822;\s*<?([\w.+-]+@[\w.-]+)>?", text, re.I): 
            addrs.append(m.group(1))

        # Pattern 1b: No space after semicolon: Final-Recipient: rfc822;email@domain.com
        for m in re.finditer(r"Final-Recipient:\s*rfc822;<?([\w.+-]+@[\w.-]+)>?", text, re.I): 
            if m.group(1) not in addrs:
                addrs.append(m.group(1))

        # Pattern 2: Generic Final-Recipient format
        for m in re.finditer(r"Final-Recipient:[^;]*;\s*<?([\w.+-]+@[\w.-]+)>?", text, re.I): 
            if m.group(1) not in addrs:
                addrs.append(m.group(1))

        # Pattern 3: Original-Recipient
        for m in re.finditer(r"Original-Recipient:\s*rfc822;\s*<?([\w.+-]+@[\w.-]+)>?", text, re.I):
            if m.group(1) not in addrs:
                addrs.append(m.group(1))

        # Pattern 4: To: field in original message
        for m in re.finditer(r"To:\s*<?([\w.+-]+@[\w.-]+)>?", text, re.I):
            if m.group(1) not in addrs:
                addrs.append(m.group(1))

        # Pattern 5: "Your message to email couldn't be delivered"
        for m in re.finditer(r"Your message to\s*<?([\w.+-]+@[\w.-]+)>?\s*couldn'?t be delivered", text, re.I):
            if m.group(1) not in addrs:
                addrs.append(m.group(1))

        # Pattern 5b: "message to <email> was undeliverable"
        for m in re.finditer(r"message to\s*<?([\w.+-]+@[\w.-]+)>?\s*was undeliverable", text, re.I):
            if m.group(1) not in addrs:
                addrs.append(m.group(1))

        # Pattern 5c: "was not delivered to" / "couldn't be delivered to"
        for m in re.finditer(r"(?:was not delivered to|wasn'?t delivered to|could not be delivered to|couldn't be delivered to|failed to deliver to)\s*<?([\w.+-]+@[\w.-]+)>?", text, re.I):
            if m.group(1) not in addrs:
                addrs.append(m.group(1))

        # Pattern 6: "Address not found" followed by email mention
        for m in re.finditer(r"Address not found.*?(?:to|for)\s*<?([\w.+-]+@[\w.-]+)>?", text, re.I | re.DOTALL):
            if m.group(1) not in addrs:
                addrs.append(m.group(1))

        # Pattern 7: Postfix/Exim format: <email> on its own line, or email: error
        for m in re.finditer(r"^\s*<([\w.+-]+@[\w.-]+)>:?\s*$", text, re.I | re.M):
            if m.group(1) not in addrs:
                addrs.append(m.group(1))

        # Pattern 7b: email followed by error description
        for m in re.finditer(r"([\w.+-]+@[\w.-]+):\s*(?:user unknown|mailbox unavailable|no such user|does not exist|mailbox full|invalid user|unknown local-part)", text, re.I):
            if m.group(1) not in addrs:
                addrs.append(m.group(1))

        # Pattern 8: "did not reach the following recipient(s):" followed by email
        for m in re.finditer(r"did not reach.*?([\w.+-]+@[\w.-]+)", text, re.I | re.DOTALL):
            if m.group(1) not in addrs:
                addrs.append(m.group(1))

        # Pattern 9: "The following address(es) failed:" followed by email
        for m in re.finditer(r"address(?:es)? failed.*?([\w.+-]+@[\w.-]+)", text, re.I | re.DOTALL):
            if m.group(1) not in addrs:
                addrs.append(m.group(1))

        # Pattern 10: Generic email in angle brackets (fallback)
        if not addrs:
            for m in re.finditer(r"<([\w.+-]+@[\w.-]+)>", text):
                if m.group(1) not in addrs:
                    addrs.append(m.group(1))

        # Pattern 11: Bare emails in bounce context (last resort - strict)
        if not addrs and any(k in text.lower() for k in ["delivery", "bounce", "failed", "undelivered", "address not found", "recipient", "mailer-daemon", "postmaster"]):
            for m in re.finditer(r"([\w.+-]+@[\w.-]+)", text):
                email = m.group(1)
                # STRICT: Skip common false positives
                if any(x in email for x in ["mailer-daemon", "postmaster", "robopirate.in", "google.com", "gmail.com", "instagram", "facebook", "twitter", "linkedin", "youtube", "2x", "3x", "1x", "wght", "size", "color", "font"]):
                    continue
                # STRICT: Must look like a real domain
                if "@" in email:
                    domain = email.split("@")[1]
                    if "." not in domain or len(domain) < 4:
                        continue
                if email not in addrs:
                    addrs.append(email)

        return addrs

    def _check_reply_scan(self, now: datetime):
        last = self.db.get_meta("last_reply_scan")
        if last and (now - datetime.fromisoformat(last)) < timedelta(minutes=REPLY_INTERVAL): return
        self.scan_replies()

    def scan_replies(self, days_back: int = 3) -> int:
        """Scan inbox for replies from recipients. Checks ALL emails (read and unread).

        Uses multiple search strategies to catch every possible reply.
        """
        after = int((datetime.now() - timedelta(days=days_back)).timestamp())

        # Search 1: ALL emails in inbox (not just unread)
        msgs_all = self.gmail.search_messages(f"in:inbox after:{after}", 200)
        self._log(f"DEBUG REPLY SCAN: {len(msgs_all)} total inbox messages found")

        # Search 2: Sent folder to find threads we started
        msgs_sent = self.gmail.search_messages(f"in:sent after:{after}", 100)
        self._log(f"DEBUG REPLY SCAN: {len(msgs_sent)} sent messages found")

        # Search 3: Any email with Re: in subject (replies)
        msgs_re = self.gmail.search_messages(f"in:inbox subject:Re: after:{after}", 100)
        self._log(f"DEBUG REPLY SCAN: {len(msgs_re)} 'Re:' messages found")

        # Combine and deduplicate
        seen_ids = set()
        all_msgs = []
        for m in msgs_all + msgs_sent + msgs_re:
            if m['id'] not in seen_ids:
                seen_ids.add(m['id'])
                all_msgs.append(m)

        self._log(f"DEBUG REPLY SCAN: {len(all_msgs)} unique messages to check")

        new_count = 0
        checked_count = 0

        for msg in all_msgs:
            from_addr = msg.get("from", "").lower()
            subject = msg.get("subject", "").lower()
            body = msg.get("body", "") or ""

            # Skip our own emails
            if "robopirate" in from_addr:
                continue

            # Skip auto-replies and bounces (handled by bounce scan)
            if self._is_auto_reply(subject, body):
                self._log(f"DEBUG REPLY: Skipping auto-reply from {from_addr}")
                continue
            if "mailer-daemon" in from_addr or "postmaster" in from_addr:
                continue

            checked_count += 1
            self._log(f"DEBUG REPLY CHECK: from={from_addr[:50]} subj={subject[:60]}")

            # Find matching recipient in our database
            rows = self.db.execute("""SELECT r.id, s.id as send_id, s.draft_id, s.day
                FROM recipients r JOIN sends s ON s.recipient_id=r.id
                WHERE r.email=? AND s.status!='replied'""", (from_addr,)).fetchall()

            if rows:
                self._log(f"DEBUG REPLY MATCH: {from_addr} matched {len(rows)} send records")
            else:
                # Also check if this email is in our recipients list at all
                check = self.db.execute("SELECT id FROM recipients WHERE email=?", (from_addr,)).fetchone()
                if check:
                    self._log(f"DEBUG REPLY: {from_addr} is in recipients but no matching send record")
                continue

            for rec_id, send_id, draft_id, day in rows:
                if self.db.execute("SELECT 1 FROM replies WHERE message_id=?", (msg["id"],)).fetchone(): 
                    self._log(f"DEBUG REPLY: Message {msg['id']} already in replies table")
                    continue
                body = msg.get("body", "")[:2000]
                self.db.execute("""INSERT INTO replies (send_id, thread_id, message_id, from_addr, subject, body, received_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?)""",
                    (send_id, msg.get("threadId", ""), msg["id"], from_addr, msg.get("subject", ""), body, datetime.now().isoformat()))
                self.db.execute("UPDATE sends SET status='replied' WHERE id=?", (send_id,))
                new_count += 1
                self._log(f"✅ New reply from {from_addr}: {msg.get('subject', '')[:60]}")
                break

        self.db.set_meta("last_reply_scan", datetime.now().isoformat())
        if new_count: 
            self._log(f"Found {new_count} new replies (checked {checked_count} messages)")
        else:
            self._log(f"No new replies found (checked {checked_count} messages from {len(all_msgs)} total in last {days_back} days)")
        return new_count

    def _check_eod(self, now: datetime):
        if now.hour != EOD_HOUR or now.minute > 5: return
        last = self.db.get_meta("last_eod_run")
        today = now.replace(hour=EOD_HOUR, minute=0, second=0, microsecond=0)
        if last and datetime.fromisoformat(last) >= today: return
        self.draft_replies_eod()

    def draft_replies_eod(self) -> dict:
        import requests
        pending = self.db.execute("SELECT * FROM replies WHERE status='pending'").fetchall()
        counts = {"positive": 0, "neutral": 0, "hostile": 0, "unsubscribe": 0, "drafted": 0}

        for row in pending:
            reply_id, send_id, thread_id, message_id, from_addr, subject, body, *_ = row
            rec = self.db.execute("""SELECT r.*, s.day, s.subject as orig_subject
                FROM recipients r JOIN sends s ON s.recipient_id=r.id WHERE s.id=?""", (send_id,)).fetchone()
            if not rec: continue

            seq_id = rec[1]
            persona = SEQUENCES.get(seq_id, {}).get("persona", "school")
            name, org = rec[3], rec[4]

            system = self._persona_prompt(persona)
            user = f"Recipient: {name} from {org}. Original: {rec[10]}. Reply: --- {body} --- Return JSON: {{sentiment, summary, draft_html}}"

            try:
                r = requests.post(f"{self.ollama_url}/api/chat", json={
                    "model": "gpt-oss:20b-cloud",
                    "messages": [{"role": "system", "content": system}, {"role": "user", "content": user}],
                    "stream": False
                }, timeout=120)
                content = r.json()["message"]["content"]
                m = re.search(r"\{.*\}", content, re.DOTALL)
                if not m: continue
                result = json.loads(m.group())

                sentiment = result.get("sentiment", "neutral")
                counts[sentiment] = counts.get(sentiment, 0) + 1

                if sentiment in ("hostile", "unsubscribe"):
                    self.db.blacklist_add(from_addr, f"sentiment:{sentiment}")
                    self.db.execute("UPDATE replies SET status='handled', sentiment=? WHERE id=?", (sentiment, reply_id))
                    self._log(f"Auto-blacklisted {from_addr} ({sentiment})")
                    continue

                draft = self.gmail.draft_reply(thread_id, result.get("draft_html", ""), f"Re: {subject}" if not subject.startswith("Re:") else subject)
                draft_id = draft.get("id") if draft else None
                self.db.execute("UPDATE replies SET status='drafted', sentiment=?, summary=?, draft_reply_id=? WHERE id=?",
                                (sentiment, result.get("summary", ""), draft_id, reply_id))
                counts["drafted"] += 1
                self._log(f"Drafted reply for {from_addr} ({sentiment}) -- waiting for your approval")
            except Exception as e:
                self._log(f"EOD draft failed: {e}")

        self.db.set_meta("last_eod_run", datetime.now().isoformat())
        self._log(f"EOD complete: {counts}")
        return counts

    def _persona_prompt(self, persona: str) -> str:
        return {
            "school": "You are the RoboPirate school outreach team. Warm, professional HTML emails to Indian private school principals. Never salesy.",
            "csr": "You are the RoboPirate CSR team. Formal, impact-focused emails to CSR heads. Data-driven and professional.",
        }.get(persona, "")

    # -- Morning Brief --
    def _check_morning_brief(self, now: datetime):
        if now.hour != MORNING_HOUR or now.minute > 5: return
        last = self.db.get_meta("last_morning_brief")
        today = now.replace(hour=0, minute=0, second=0, microsecond=0)
        if last and datetime.fromisoformat(last) >= today: return

        brief = self.morning_brief()
        if self.brief_email:
            try:
                self.gmail.send_email(self.brief_email, f"Raj Brief -- {now.strftime('%d %b %Y')}", brief.replace("\n", "<br>"))
                self._log("Morning brief sent")
            except Exception as e:
                self._log(f"Brief failed: {e}")
        self.db.set_meta("last_morning_brief", now.isoformat())

    def morning_brief(self) -> str:
        today = datetime.now().strftime("%d %b %Y")
        yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
        lines = ["=" * 40, f"RAJ BRIEF -- {today}", "=" * 40, "YESTERDAY"]

        for seq_id in SEQUENCES:
            stats = self.db.execute("SELECT day, COUNT(*) FROM sends WHERE recipient_id IN (SELECT id FROM recipients WHERE sequence_id=?) AND DATE(created_at)=? GROUP BY day", (seq_id, yesterday)).fetchall()
            if stats:
                for day, count in stats: lines.append(f"  {seq_id.upper()} Day {day}: {count} sent")
            else: lines.append(f"  {seq_id.upper()}: No batches")

        replies = self.db.execute("SELECT sentiment, COUNT(*) FROM replies WHERE DATE(received_at)=? OR status IN ('pending','drafted') GROUP BY sentiment", (yesterday,)).fetchall()
        rc = {k: 0 for k in ["positive", "neutral", "hostile", "unsubscribe"]}
        for s, c in replies:
            if s in rc: rc[s] = c
        lines.extend(["", f"REPLIES ({sum(rc.values())} total)", f"  -- {rc['positive']} positive", f"  -- {rc['neutral']} neutral", f"  -- {rc['hostile'] + rc['unsubscribe']} hostile -- blacklisted", "  -> Review drafts in Gmail before sending"])

        bounces = self.db.execute("SELECT email, reason FROM blacklist WHERE DATE(added_at)=? OR reason LIKE 'bounce %'", (yesterday,)).fetchall()
        lines.extend(["", f"BOUNCES ({len(bounces)} overnight)"])
        for email, reason in bounces[:5]: lines.append(f"  {email} -- {reason}")

        lines.extend(["", "DUE TODAY"])
        for seq_id in SEQUENCES:
            for day in SEQUENCES[seq_id]["days"]:
                due = len(self.due_recipients(seq_id, day))
                if due: lines.append(f"  {seq_id.upper()} Day {day}: {due} recipients")

        lines.extend(["", "YOUR ACTIONS", "  1. Review reply drafts in Gmail (DRAFT-ONLY for approval)", "  2. Sequences auto-send at 10 AM -- no action needed", "  3. Reply STOP SCHOOL / STOP CSR / STOP ALL to pause", "=" * 40])
        return "\n".join(lines)

    # -- Emergency Commands --
    def _check_emergency_commands(self, now: datetime):
        last = self.db.get_meta("last_emergency_scan")
        if last and (now - datetime.fromisoformat(last)) < timedelta(minutes=EMERGENCY_INTERVAL): return

        after = int((datetime.now() - timedelta(hours=1)).timestamp())
        msgs = self.gmail.search_messages(f"in:inbox from:me subject:(STOP SCHOOL OR STOP CSR OR STOP ALL OR RESUME) after:{after}", 10)

        for msg in msgs:
            subj = msg.get("subject", "").upper()
            if "STOP SCHOOL" in subj: self.db.set_meta("pause_school", "true"); self._log("SCHOOL paused")
            elif "STOP CSR" in subj: self.db.set_meta("pause_csr", "true"); self._log("CSR paused")
            elif "STOP ALL" in subj: self.pause(); self._log("ALL paused")
            elif "RESUME" in subj: self.resume(); self.db.execute("DELETE FROM meta WHERE key LIKE 'pause_%'"); self._log("All resumed")

        self.db.set_meta("last_emergency_scan", now.isoformat())

    # -- Campaign State Export --
    def export_campaign_state(self) -> str:
        from pathlib import Path
        now = datetime.now().strftime("%d %b %Y %H:%M")
        lines = [
            f"# Raj Campaign State -- {now}",
            "",
            "## Sequences",
            ""
        ]

        for seq_id in SEQUENCES:
            lines.append(f"### {seq_id.upper()}")
            cfg = SEQUENCES[seq_id]
            for day in cfg["days"]:
                due = self.due_recipients(seq_id, day)
                sent = self.db.execute(
                    "SELECT COUNT(DISTINCT recipient_id) FROM sends WHERE day=? AND status IN ('sent','drafted') AND recipient_id IN (SELECT id FROM recipients WHERE sequence_id=?)",
                    (day, seq_id)
                ).fetchone()[0]
                total = self.db.execute("SELECT COUNT(*) FROM recipients WHERE sequence_id=?", (seq_id,)).fetchone()[0]
                lines.append(f"- Day {day}: {sent}/{total} sent | {len(due)} due")
            lines.append("")

        pending = self.db.execute("SELECT sequence_id, day, COUNT(*) FROM pending_resumes WHERE status='pending' GROUP BY sequence_id, day").fetchall()
        if pending:
            lines.append("## Pending Resumes (Quota Interruptions)")
            for seq_id, day, count in pending:
                lines.append(f"- {seq_id.upper()} Day {day}: {count} emails waiting to resume")
            lines.append("")
        else:
            lines.append("## Pending Resumes")
            lines.append("- None. All batches completed cleanly.")
            lines.append("")

        pending_replies = self.db.execute("SELECT COUNT(*) FROM replies WHERE status='pending'").fetchone()[0]
        drafted_replies = self.db.execute("SELECT COUNT(*) FROM replies WHERE status='drafted'").fetchone()[0]
        lines.append("## Replies")
        lines.append(f"- Pending: {pending_replies}")
        lines.append(f"- Drafted (awaiting approval): {drafted_replies}")
        lines.append("")

        bl_count = self.db.execute("SELECT COUNT(*) FROM blacklist").fetchone()[0]
        bl_recent = self.db.execute("SELECT email, reason FROM blacklist ORDER BY added_at DESC LIMIT 10").fetchall()
        lines.append(f"## Blacklist ({bl_count} total)")
        for email, reason in bl_recent:
            lines.append(f"- `{email}` -- {reason}")
        lines.append("")

        lines.append("## Engine Status")
        lines.append(f"- Running: {self.is_running()}")
        lines.append(f"- Paused: {self.is_paused()}")
        lines.append(f"- Last bounce scan: {self.db.get_meta('last_bounce_scan') or 'Never'}")
        lines.append(f"- Last reply scan: {self.db.get_meta('last_reply_scan') or 'Never'}")
        lines.append("")

        lines.append("---")
        lines.append("*Auto-generated by Raj Campaign Engine*")

        md = "\n".join(lines)

        state_path = Path(__file__).parent / "campaign_state.md"
        with open(state_path, "w", encoding="utf-8") as f:
            f.write(md)

        self._log(f"Campaign state exported to {state_path}")
        return md

    # -- Quota Rollback & Resume --
    def resume_batch(self, seq_id: str, day: int, limit=None) -> BatchResult:
        pending = self.db.execute(
            "SELECT recipient_id, subject FROM pending_resumes WHERE sequence_id=? AND day=? AND status='pending' ORDER BY id",
            (seq_id, day)
        ).fetchall()

        if not pending:
            self._log(f"No pending resumes for {seq_id.upper()} Day {day}")
            return BatchResult(queued=0, sent=0)

        if limit:
            pending = pending[:limit]

        self._log(f"Resuming {seq_id.upper()} Day {day}: {len(pending)} pending")
        sent = 0

        for rec_id, subject in pending:
            rec_row = self.db.execute("SELECT * FROM recipients WHERE id=?", (rec_id,)).fetchone()
            if not rec_row:
                continue
            rec = Recipient(*rec_row)

            subj, body = self.render(seq_id, day, rec)
            if not subj:
                subj = subject

            try:
                msg = self.gmail.send_email(rec.email, subj, body)
                self.db.campaign_queue_send(rec.id, day, subj, msg.get("id"), "sent")
                self.db.execute(
                    "UPDATE pending_resumes SET status='sent', resumed_at=? WHERE recipient_id=? AND sequence_id=? AND day=? AND status='pending'",
                    (datetime.now().isoformat(), rec.id, seq_id, day)
                )
                sent += 1
                self._log(f"Resumed send to {rec.email}")
                time.sleep(SEND_DELAY)
            except Exception as e:
                err = str(e)
                if "quota" in err.lower() or "rate" in err.lower() or "limit" in err.lower():
                    self._log("Rate limit hit again during resume. Stopping.")
                    break
                self._log(f"Resume failed for {rec.email}: {e}")
                self.db.execute(
                    "UPDATE pending_resumes SET status='error', error=? WHERE recipient_id=? AND sequence_id=? AND day=? AND status='pending'",
                    (str(e)[:200], rec.id, seq_id, day)
                )

        self.db.commit()
        self._log(f"Resume complete: {sent}/{len(pending)} sent")
        return BatchResult(queued=len(pending), sent=sent)

    def backdate_sequence(self, seq_id: str, day: int, days_ago: int) -> int:
        cutoff = (datetime.now() - timedelta(days=days_ago)).isoformat()
        rows = self.db.execute(
            "SELECT id, created_at FROM sends WHERE recipient_id IN (SELECT id FROM recipients WHERE sequence_id=?) AND day=? AND created_at > ?",
            (seq_id, day, cutoff)
        ).fetchall()

        count = 0
        for send_id, created_at in rows:
            new_time = (datetime.fromisoformat(created_at) - timedelta(days=days_ago)).isoformat()
            self.db.execute("UPDATE sends SET created_at=? WHERE id=?", (new_time, send_id))
            count += 1

        self.db.commit()
        self._log(f"Backdated {count} sends for {seq_id.upper()} Day {day} by {days_ago} days")
        return count

    def import_blacklist_file(self, filepath: str) -> int:
        from pathlib import Path
        path = Path(filepath)
        if not path.exists():
            self._log(f"Blacklist file not found: {filepath}")
            return 0

        emails = []
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                found = re.findall(r"[\w.+-]+@[\w.-]+", line)
                emails.extend(found)

        unique = list(set(e.lower().strip() for e in emails if "@" in e))
        count = 0
        for email in unique:
            if not self.db.blacklist_has(email):
                self.db.blacklist_add(email, f"imported_from_file {path.name}")
                count += 1

        self._log(f"Imported {count} new blacklisted emails from {path.name} ({len(unique)} found)")
        return count
