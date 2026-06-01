"""
smart_importer.py — Intelligent Lead Import System for Raj v1.0
Reads Excel, CSV, TXT files and auto-detects columns like Brevo does.
Auto-creates batches of configurable size (50/100/150) with auto-advance.
"""

import re
import csv
import json
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass

@dataclass
class ColumnMapping:
    email: str = None
    name: str = None
    org: str = None
    phone: str = None
    city: str = None
    state: str = None
    country: str = None
    designation: str = None
    website: str = None
    extra: Dict = None

class SmartImporter:
    """Intelligent file reader that auto-detects columns from any format."""

    # Column name patterns for auto-detection
    EMAIL_PATTERNS = [
        r'^e[-_\s]?mail', r'^email[-_\s]?id', r'^email[-_\s]?address', r'^mail[-_\s]?id',
        r'^contact[-_\s]?email', r'^email[-_\s]?1', r'^primary[-_\s]?email',
        r'^work[-_\s]?email', r'^business[-_\s]?email', r'^official[-_\s]?email'
    ]

    NAME_PATTERNS = [
        r'^name$', r'^full[-_\s]?name', r'^contact[-_\s]?name', r'^person[-_\s]?name',
        r'^first[-_\s]?name', r'^last[-_\s]?name', r'^principal[-_\s]?name',
        r'^contact[-_\s]?person', r'^representative', r'^head[-_\s]?name',
        r'^owner[-_\s]?name', r'^manager[-_\s]?name', r'^director[-_\s]?name'
    ]

    ORG_PATTERNS = [
        r'^org', r'^organization', r'^company', r'^school', r'^institution',
        r'^company[-_\s]?name', r'^school[-_\s]?name', r'^institute[-_\s]?name',
        r'^org[-_\s]?name', r'^business[-_\s]?name', r'^firm[-_\s]?name',
        r'^establishment', r'^entity', r'^corporate[-_\s]?name'
    ]

    PHONE_PATTERNS = [
        r'^phone', r'^mobile', r'^contact[-_\s]?number', r'^tel', r'^telephone',
        r'^cell', r'^cellphone', r'^whatsapp', r'^phone[-_\s]?number', r'^number$'
    ]

    CITY_PATTERNS = [
        r'^city', r'^town', r'^location', r'^place', r'^area'
    ]

    STATE_PATTERNS = [
        r'^state', r'^province', r'^region', r'^district'
    ]

    COUNTRY_PATTERNS = [
        r'^country', r'^nation', r'^region'
    ]

    DESIGNATION_PATTERNS = [
        r'^designation', r'^title', r'^position', r'^role', r'^job[-_\s]?title',
        r'^post', r'^rank', r'^status'
    ]

    WEBSITE_PATTERNS = [
        r'^website', r'^url', r'^web', r'^site', r'^web[-_\s]?address',
        r'^domain', r'^link'
    ]

    def __init__(self, db, engine):
        self.db = db
        self.engine = engine
        self.mapping = ColumnMapping()
        self.detected_headers = []
        self.confidence_scores = {}

    def _normalize_header(self, header: str) -> str:
        """Normalize header for pattern matching."""
        if not header:
            return ""
        return header.lower().strip().replace(" ", "_").replace("-", "_")

    def _match_pattern(self, header: str, patterns: List[str]) -> bool:
        """Check if header matches any pattern."""
        norm = self._normalize_header(header)
        for pattern in patterns:
            if re.search(pattern, norm, re.IGNORECASE):
                return True
        return False

    def _detect_columns(self, headers: List[str]) -> ColumnMapping:
        """Auto-detect column mappings from headers."""
        mapping = ColumnMapping()
        mapping.extra = {}
        self.confidence_scores = {}

        for i, header in enumerate(headers):
            if not header:
                continue

            norm = self._normalize_header(header)

            # Email (highest priority)
            if self._match_pattern(header, self.EMAIL_PATTERNS):
                if not mapping.email:
                    mapping.email = header
                    self.confidence_scores["email"] = (header, "high")
                continue

            # Name
            if self._match_pattern(header, self.NAME_PATTERNS):
                if not mapping.name:
                    mapping.name = header
                    self.confidence_scores["name"] = (header, "high")
                continue

            # Organization
            if self._match_pattern(header, self.ORG_PATTERNS):
                if not mapping.org:
                    mapping.org = header
                    self.confidence_scores["org"] = (header, "high")
                continue

            # Phone
            if self._match_pattern(header, self.PHONE_PATTERNS):
                if not mapping.phone:
                    mapping.phone = header
                    self.confidence_scores["phone"] = (header, "medium")
                continue

            # City
            if self._match_pattern(header, self.CITY_PATTERNS):
                if not mapping.city:
                    mapping.city = header
                    self.confidence_scores["city"] = (header, "medium")
                continue

            # State
            if self._match_pattern(header, self.STATE_PATTERNS):
                if not mapping.state:
                    mapping.state = header
                    self.confidence_scores["state"] = (header, "medium")
                continue

            # Country
            if self._match_pattern(header, self.COUNTRY_PATTERNS):
                if not mapping.country:
                    mapping.country = header
                    self.confidence_scores["country"] = "low"
                continue

            # Designation
            if self._match_pattern(header, self.DESIGNATION_PATTERNS):
                if not mapping.designation:
                    mapping.designation = header
                    self.confidence_scores["designation"] = (header, "medium")
                continue

            # Website
            if self._match_pattern(header, self.WEBSITE_PATTERNS):
                if not mapping.website:
                    mapping.website = header
                    self.confidence_scores["website"] = (header, "low")
                continue

            # Everything else goes to extra
            mapping.extra[header] = None

        # Fallback: try to detect email by content if no pattern matched
        if not mapping.email:
            for header in headers:
                if "@" in str(header).lower() or ".com" in str(header).lower():
                    mapping.email = header
                    self.confidence_scores["email"] = (header, "low (content-based)")
                    break

        # Fallback: if no name column, look for anything with "name" in it
        if not mapping.name:
            for header in headers:
                if "name" in str(header).lower():
                    mapping.name = header
                    self.confidence_scores["name"] = (header, "low (fallback)")
                    break

        # Fallback: if no org column, look for anything with "school", "company", etc.
        if not mapping.org:
            for header in headers:
                h = str(header).lower()
                if any(k in h for k in ["school", "company", "org", "institute", "college", "firm"]):
                    mapping.org = header
                    self.confidence_scores["org"] = (header, "low (fallback)")
                    break

        self.mapping = mapping
        return mapping

    def _read_file(self, filepath: str) -> Tuple[List[str], List[Dict]]:
        """Read any file type and return (headers, rows_as_dicts)."""
        path = Path(filepath)
        ext = path.suffix.lower()

        if ext in [".xlsx", ".xls"]:
            return self._read_excel(filepath)
        elif ext == ".csv":
            return self._read_csv(filepath)
        elif ext in [".txt", ".tsv", ".dat"]:
            return self._read_delimited(filepath)
        else:
            raise ValueError(f"Unsupported file type: {ext}. Use .xlsx, .xls, .csv, .txt, .tsv")

    def _read_excel(self, filepath: str) -> Tuple[List[str], List[Dict]]:
        """Read Excel file."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl")

        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        # Get headers from first row
        headers = []
        for cell in ws[1]:
            val = cell.value
            headers.append(str(val).strip() if val is not None else "")

        # Get data rows
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = val if val is not None else ""
            rows.append(row_dict)

        return headers, rows

    def _read_csv(self, filepath: str) -> Tuple[List[str], List[Dict]]:
        """Read CSV file with auto-detected delimiter."""
        with open(filepath, 'r', encoding='utf-8-sig', errors='replace') as f:
            # Try to detect delimiter from first line
            sample = f.read(4096)
            f.seek(0)

            # Count common delimiters
            delimiters = [',', ';', '\t', '|']
            counts = {d: sample.count(d) for d in delimiters}
            delimiter = max(counts, key=counts.get)

            if counts[delimiter] < 2:
                delimiter = ','  # Default

            reader = csv.DictReader(f, delimiter=delimiter)
            headers = reader.fieldnames or []
            rows = list(reader)
            # Convert empty strings to None for consistency
            for row in rows:
                for k in row:
                    if row[k] == '':
                        row[k] = None
            return headers, rows

    def _read_delimited(self, filepath: str) -> Tuple[List[str], List[Dict]]:
        """Read tab/space delimited text file."""
        with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
            lines = [l.strip() for l in f.readlines() if l.strip()]

        if not lines:
            return [], []

        # Try tab first, then comma, then space
        for delimiter in ['\t', ',', ';']:
            headers = [h.strip() for h in lines[0].split(delimiter)]
            if len(headers) > 1:
                rows = []
                for line in lines[1:]:
                    values = [v.strip() for v in line.split(delimiter)]
                    row_dict = {}
                    for i, h in enumerate(headers):
                        row_dict[h] = values[i] if i < len(values) else ""
                    rows.append(row_dict)
                return headers, rows

        # Single column - treat as email list
        headers = ["email"]
        rows = [{"email": line} for line in lines]
        return headers, rows

    def _extract_email(self, value) -> Optional[str]:
        """Extract valid email from a value."""
        if not value:
            return None
        text = str(value).strip().lower()

        # Direct email
        if re.match(r'^[\w.+-]+@[\w.-]+\.[a-zA-Z]{2,}$', text):
            return text

        # Extract email from text
        match = re.search(r'[\w.+-]+@[\w.-]+\.[a-zA-Z]{2,}', text)
        if match:
            return match.group(0)

        return None

    def _clean_value(self, value) -> str:
        """Clean a value for storage."""
        if value is None:
            return ""
        text = str(value).strip()
        # Remove common noise
        text = re.sub(r'\s+', ' ', text)
        return text

    def analyze_file(self, filepath: str) -> Dict:
        """Analyze a file and return detection report."""
        headers, rows = self._read_file(filepath)
        mapping = self._detect_columns(headers)

        # Sample first 3 rows for preview
        sample = []
        for row in rows[:3]:
            sample.append({h: row.get(h, "") for h in headers})

        # Count valid emails
        valid_emails = 0
        invalid_emails = 0
        if mapping.email:
            for row in rows:
                email = self._extract_email(row.get(mapping.email))
                if email:
                    valid_emails += 1
                else:
                    invalid_emails += 1

        return {
            "filename": Path(filepath).name,
            "total_rows": len(rows),
            "headers": headers,
            "mapping": {
                "email": mapping.email,
                "name": mapping.name,
                "org": mapping.org,
                "phone": mapping.phone,
                "city": mapping.city,
                "state": mapping.state,
                "country": mapping.country,
                "designation": mapping.designation,
                "website": mapping.website,
            },
            "confidence": self.confidence_scores,
            "sample_rows": sample,
            "valid_emails": valid_emails,
            "invalid_emails": invalid_emails,
            "ready_to_import": mapping.email is not None and valid_emails > 0
        }

    def import_leads(self, filepath: str, sequence_id: str, batch_size: int = 50,
                     custom_mapping: Dict = None, auto_create_batches: bool = True,
                     batch_name_prefix: str = None, start_day: int = 1) -> Dict:
        """
        Import leads from file and optionally auto-create batches.

        Args:
            filepath: Path to Excel/CSV/TXT file
            sequence_id: "school" or "csr"
            batch_size: Number of leads per batch (50, 100, 150, etc.)
            custom_mapping: Override auto-detected column mapping
            auto_create_batches: Create batches automatically
            batch_name_prefix: Prefix for batch names (e.g., "Pune-Schools")
            start_day: Starting day offset (1, 3, 5, 7, 10)

        Returns:
            Dict with import results and batch info
        """
        headers, rows = self._read_file(filepath)

        # Use custom mapping if provided, else auto-detect
        if custom_mapping:
            self.mapping = ColumnMapping(**custom_mapping)
        else:
            self._detect_columns(headers)

        mapping = self.mapping

        if not mapping.email:
            return {
                "success": False,
                "error": "Could not detect email column. Please provide custom mapping.",
                "detected_headers": headers
            }

        # Import recipients
        imported = 0
        skipped = 0
        blacklisted = 0
        recipient_ids = []

        for row in rows:
            email = self._extract_email(row.get(mapping.email))
            if not email:
                skipped += 1
                continue

            # Check blacklist
            if self.db.blacklist_has(email):
                blacklisted += 1
                skipped += 1
                continue

            name = self._clean_value(row.get(mapping.name, "Principal"))
            if not name or name == "":
                name = "Principal"

            org = self._clean_value(row.get(mapping.org, ""))

            # Build extra JSON with all other columns
            extra = {}
            for h in headers:
                if h not in [mapping.email, mapping.name, mapping.org]:
                    val = row.get(h)
                    if val is not None and str(val).strip():
                        extra[h] = self._clean_value(val)

            # Add phone, city, state, etc. to extra if detected
            if mapping.phone and row.get(mapping.phone):
                extra["phone"] = self._clean_value(row.get(mapping.phone))
            if mapping.city and row.get(mapping.city):
                extra["city"] = self._clean_value(row.get(mapping.city))
            if mapping.state and row.get(mapping.state):
                extra["state"] = self._clean_value(row.get(mapping.state))
            if mapping.country and row.get(mapping.country):
                extra["country"] = self._clean_value(row.get(mapping.country))
            if mapping.designation and row.get(mapping.designation):
                extra["designation"] = self._clean_value(row.get(mapping.designation))
            if mapping.website and row.get(mapping.website):
                extra["website"] = self._clean_value(row.get(mapping.website))

            extra_json = json.dumps(extra) if extra else None

            ok, err = self.db.recipient_add(sequence_id, email, name, org, extra_json)
            if ok:
                imported += 1
                # Get the recipient ID
                rid_row = self.db.execute(
                    "SELECT id FROM recipients WHERE sequence_id=? AND email=?",
                    (sequence_id, email)
                ).fetchone()
                if rid_row:
                    recipient_ids.append(rid_row[0])
            else:
                skipped += 1

        self.db.commit()

        result = {
            "success": True,
            "imported": imported,
            "skipped": skipped,
            "blacklisted": blacklisted,
            "total_rows": len(rows),
            "sequence": sequence_id,
            "mapping": {
                "email": mapping.email,
                "name": mapping.name,
                "org": mapping.org,
                "phone": mapping.phone,
                "city": mapping.city,
                "state": mapping.state,
                "designation": mapping.designation,
                "website": mapping.website,
            }
        }

        # Auto-create batches
        batches_created = []
        if auto_create_batches and imported > 0:
            batches_created = self._create_batches(
                recipient_ids, sequence_id, batch_size,
                batch_name_prefix or Path(filepath).stem, start_day
            )
            result["batches"] = batches_created

        # Log action
        self.db.log_action("smart_import",
            f"File: {Path(filepath).name}, Seq: {sequence_id}, Imported: {imported}, "
            f"Skipped: {skipped}, Batches: {len(batches_created)}", "user")

        return result

    def import_to_pool(self, filepath: str, sequence_id: str,
                       custom_mapping: Dict = None) -> Dict:
        """
        Import leads to POOL only (no batch creation).
        All leads go to DB first as single source of truth.
        User creates batches later from the pool.

        Args:
            filepath: Path to Excel/CSV/TXT file
            sequence_id: "school" or "csr"
            custom_mapping: Override auto-detected column mapping

        Returns:
            Dict with import results and pool stats
        """
        headers, rows = self._read_file(filepath)

        if custom_mapping:
            self.mapping = ColumnMapping(**custom_mapping)
        else:
            self._detect_columns(headers)

        mapping = self.mapping

        if not mapping.email:
            return {
                "success": False,
                "error": "Could not detect email column. Please provide custom mapping.",
                "detected_headers": headers
            }

        imported = 0
        skipped = 0
        blacklisted = 0
        duplicates = 0

        for row in rows:
            email = self._extract_email(row.get(mapping.email))
            if not email:
                skipped += 1
                continue

            if self.db.blacklist_has(email):
                blacklisted += 1
                skipped += 1
                continue

            name = self._clean_value(row.get(mapping.name, "Principal"))
            if not name or name == "":
                name = "Principal"

            org = self._clean_value(row.get(mapping.org, ""))

            extra = {}
            for h in headers:
                if h not in [mapping.email, mapping.name, mapping.org]:
                    val = row.get(h)
                    if val is not None and str(val).strip():
                        extra[h] = self._clean_value(val)

            if mapping.phone and row.get(mapping.phone):
                extra["phone"] = self._clean_value(row.get(mapping.phone))
            if mapping.city and row.get(mapping.city):
                extra["city"] = self._clean_value(row.get(mapping.city))
            if mapping.state and row.get(mapping.state):
                extra["state"] = self._clean_value(row.get(mapping.state))
            if mapping.country and row.get(mapping.country):
                extra["country"] = self._clean_value(row.get(mapping.country))
            if mapping.designation and row.get(mapping.designation):
                extra["designation"] = self._clean_value(row.get(mapping.designation))
            if mapping.website and row.get(mapping.website):
                extra["website"] = self._clean_value(row.get(mapping.website))

            extra_json = json.dumps(extra) if extra else None

            ok, err = self.db.recipient_add(sequence_id, email, name, org, extra_json)
            if ok:
                imported += 1
            else:
                if "UNIQUE" in str(err).upper() or "duplicate" in str(err).lower():
                    duplicates += 1
                skipped += 1

        self.db.commit()

        pool_count = self.db.get_pool_count(sequence_id)
        total_count = self.db.recipient_count(sequence_id)

        self.db.log_action("import_to_pool",
            f"File: {Path(filepath).name}, Seq: {sequence_id}, Imported: {imported}, "
            f"Skipped: {skipped}, Blacklisted: {blacklisted}, Duplicates: {duplicates}, "
            f"Pool: {pool_count}, Total: {total_count}", "user")

        return {
            "success": True,
            "imported": imported,
            "skipped": skipped,
            "blacklisted": blacklisted,
            "duplicates": duplicates,
            "total_rows": len(rows),
            "sequence": sequence_id,
            "pool_count": pool_count,
            "total_in_sequence": total_count,
            "mapping": {
                "email": mapping.email,
                "name": mapping.name,
                "org": mapping.org,
                "phone": mapping.phone,
                "city": mapping.city,
                "state": mapping.state,
                "designation": mapping.designation,
                "website": mapping.website,
            }
        }

    def _create_batches(self, recipient_ids: List[int], sequence_id: str,
                        batch_size: int, name_prefix: str, start_day: int) -> List[Dict]:
        """Create batches from recipient IDs with auto-scheduling."""
        batches = []

        # Chunk recipients into batches
        for i in range(0, len(recipient_ids), batch_size):
            chunk = recipient_ids[i:i + batch_size]
            batch_num = (i // batch_size) + 1
            batch_name = f"{name_prefix}-B{batch_num}"

            # Schedule: first batch immediate, subsequent batches +2 days each
            if i == 0:
                scheduled = datetime.now().isoformat()
                status = "draft"  # User starts manually
            else:
                days_offset = (i // batch_size) * 2
                scheduled = (datetime.now() + timedelta(days=days_offset)).replace(
                    hour=10, minute=0, second=0, microsecond=0
                ).isoformat()
                status = "scheduled"

            batch_id = self.db.batch_create(
                batch_name, sequence_id, scheduled_at=scheduled,
                stagger_minutes=2, day_offset=start_day
            )

            # Add recipients to batch
            for rid in chunk:
                self.db.batch_add_recipient(batch_id, rid)

            batches.append({
                "id": batch_id,
                "name": batch_name,
                "recipients": len(chunk),
                "scheduled": scheduled,
                "status": status,
                "day_offset": start_day
            })

        return batches

    def get_import_preview(self, filepath: str, max_rows: int = 5) -> Dict:
        """Get a preview of how the file will be imported."""
        headers, rows = self._read_file(filepath)
        mapping = self._detect_columns(headers)

        preview_rows = []
        for row in rows[:max_rows]:
            preview = {
                "email": self._extract_email(row.get(mapping.email)),
                "name": self._clean_value(row.get(mapping.name, "")),
                "org": self._clean_value(row.get(mapping.org, "")),
                "phone": self._clean_value(row.get(mapping.phone, "")) if mapping.phone else "",
                "city": self._clean_value(row.get(mapping.city, "")) if mapping.city else "",
                "extra_fields": {h: self._clean_value(v) for h, v in row.items()
                               if h not in [mapping.email, mapping.name, mapping.org, mapping.phone, mapping.city]
                               and v is not None and str(v).strip()}
            }
            preview_rows.append(preview)

        return {
            "filename": Path(filepath).name,
            "total_rows": len(rows),
            "detected_mapping": {
                "email": mapping.email,
                "name": mapping.name,
                "org": mapping.org,
                "phone": mapping.phone,
                "city": mapping.city,
                "state": mapping.state,
                "designation": mapping.designation,
                "website": mapping.website,
            },
            "confidence": self.confidence_scores,
            "preview": preview_rows,
            "unmapped_columns": [h for h in headers if h not in [
                mapping.email, mapping.name, mapping.org, mapping.phone,
                mapping.city, mapping.state, mapping.country,
                mapping.designation, mapping.website
            ] and h]
        }
